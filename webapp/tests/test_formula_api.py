import os
import tempfile
import unittest
from pathlib import Path

import openpyxl
from werkzeug.security import generate_password_hash

os.environ["SECRET_KEY"] = "test-secret-key"
os.environ["PASSWORD_HASH"] = generate_password_hash("secret123")
os.environ["LOGIN_RATE_LIMIT_BURST"] = "1000 per hour"
os.environ["LOGIN_RATE_LIMIT_SUSTAINED"] = "1000 per hour"
os.environ["DEFAULT_RATE_LIMIT"] = "10000 per hour"
os.environ.setdefault("DATA_DIR", tempfile.mkdtemp(prefix="ct_formula_api_"))

import config
import data_layer
import settings
from app import create_app


class FormulaApiTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config["TESTING"] = True
        cls.client = cls.app.test_client()

    def setUp(self):
        config.DATA_DIR = Path(tempfile.mkdtemp(prefix="ct_formula_api_case_"))
        config.BAK_DIR = config.DATA_DIR / ".bak"
        config.AUDIT_LOG = config.DATA_DIR / "audit.log"
        settings._cache = None
        self.build_workbook()
        token = self.client.get("/api/session").get_json()["csrf_token"]
        response = self.client.post(
            "/api/login", json={"password": "secret123"}, headers={"X-CSRF-Token": token}
        )
        self.assertEqual(response.status_code, 200)

    def build_workbook(self, name="f.xlsx"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for i, header in enumerate(
            ["DATE", "CUSTOMER NAME", "BANK NO.", "ACCOUNT NO.", "AMOUNT", "PAID TO", "TOTAL"], 1
        ):
            ws.cell(row=1, column=i, value=header)
        ws.cell(row=2, column=7, value="=E2*1.18")
        ws.cell(row=3, column=1, value="2026-08-01")
        ws.cell(row=3, column=2, value="manu")
        ws.cell(row=3, column=5, value=500)
        ws.cell(row=3, column=7, value=590.0)
        ws.cell(row=4, column=1, value="2026-08-02")
        ws.cell(row=4, column=2, value="asha")
        ws.cell(row=4, column=5, value=1000)
        ws.cell(row=4, column=7, value=1180.0)
        wb.save(config.DATA_DIR / name)

    def _csrf(self):
        return self.client.get("/api/session").get_json()["csrf_token"]

    def test_get_formulas_returns_all_columns(self):
        response = self.client.get("/api/formulas?wb=f.xlsx&sheet=Sheet1")
        self.assertEqual(response.status_code, 200)
        formulas = response.get_json()["formulas"]
        self.assertEqual(
            set(formulas.keys()),
            {"DATE", "CUSTOMER NAME", "BANK NO.", "ACCOUNT NO.", "AMOUNT", "PAID TO", "TOTAL"},
        )
        self.assertEqual(formulas["TOTAL"]["formula"], "=E2*1.18")
        self.assertEqual(formulas["TOTAL"]["ref"], "G")
        self.assertEqual(formulas["AMOUNT"]["formula"], "")
        self.assertEqual(formulas["AMOUNT"]["ref"], "E")
        self.assertEqual(formulas["DATE"]["formula"], "")
        self.assertEqual(formulas["DATE"]["ref"], "A")

    def test_put_updates_existing_formula_and_recomputes(self):
        put = self.client.put(
            "/api/formulas?wb=f.xlsx&sheet=Sheet1",
            json={"formulas": {"TOTAL": "=E2*2"}},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(put.status_code, 200)
        self.assertEqual(put.get_json()["formulas"]["TOTAL"], "=E2*2")
        self.assertEqual(put.get_json()["warnings"], [])

        sheet_data = self.client.get("/api/sheet-data?wb=f.xlsx&sheet=Sheet1").get_json()
        by_row = {r["excel_row"]: r["values"] for r in sheet_data["rows"]}
        self.assertEqual(by_row[3]["TOTAL"], "1000")
        self.assertEqual(by_row[4]["TOTAL"], "2000")

        add = self.client.post(
            "/api/rows",
            json={"wb": "f.xlsx", "sheet": "Sheet1",
                  "values": {"DATE": "2026-08-03", "CUSTOMER NAME": "gst", "AMOUNT": "200"}},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(add.status_code, 200)
        self.assertEqual(add.get_json()["row"]["values"]["TOTAL"], "400")

    def test_put_adds_new_formula_column(self):
        put = self.client.put(
            "/api/formulas?wb=f.xlsx&sheet=Sheet1",
            json={"formulas": {"GST": "=E2*0.05"}},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(put.status_code, 200)
        self.assertEqual(put.get_json()["formulas"]["GST"], "=E2*0.05")

        sheet_data = self.client.get("/api/sheet-data?wb=f.xlsx&sheet=Sheet1").get_json()
        self.assertIn("GST", sheet_data["headers"])
        self.assertIn("GST", sheet_data["formula_cols"])
        by_row = {r["excel_row"]: r["values"] for r in sheet_data["rows"]}
        self.assertEqual(by_row[3]["GST"], "25")
        self.assertEqual(by_row[4]["GST"], "50")

        add = self.client.post(
            "/api/rows",
            json={"wb": "f.xlsx", "sheet": "Sheet1",
                  "values": {"DATE": "2026-08-04", "CUSTOMER NAME": "newcol", "AMOUNT": "1000"}},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(add.status_code, 200)
        self.assertEqual(add.get_json()["row"]["values"]["GST"], "50")

    def test_put_rejects_invalid_formula_and_writes_nothing(self):
        put = self.client.put(
            "/api/formulas?wb=f.xlsx&sheet=Sheet1",
            json={"formulas": {"TOTAL": "=SUM("}},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(put.status_code, 400)
        body = put.get_json()
        self.assertIn("TOTAL", body["invalid"])
        self.assertTrue(body["invalid"]["TOTAL"])

        formulas = self.client.get("/api/formulas?wb=f.xlsx&sheet=Sheet1").get_json()["formulas"]
        self.assertEqual(formulas["TOTAL"]["formula"], "=E2*1.18")

    def test_put_invalid_is_atomic_across_columns(self):
        put = self.client.put(
            "/api/formulas?wb=f.xlsx&sheet=Sheet1",
            json={"formulas": {"TOTAL": "=E2*3", "BROKEN": "=__import__('os')"}},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(put.status_code, 400)
        self.assertIn("BROKEN", put.get_json()["invalid"])

        formulas = self.client.get("/api/formulas?wb=f.xlsx&sheet=Sheet1").get_json()["formulas"]
        self.assertEqual(formulas["TOTAL"]["formula"], "=E2*1.18")
        self.assertNotIn("BROKEN", formulas)

    def test_put_rejects_missing_column_reference(self):
        put = self.client.put(
            "/api/formulas?wb=f.xlsx&sheet=Sheet1",
            json={"formulas": {"TOTAL": "=ZZ2*2"}},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(put.status_code, 400)
        self.assertIn("TOTAL", put.get_json()["invalid"])


if __name__ == "__main__":
    unittest.main(verbosity=2)