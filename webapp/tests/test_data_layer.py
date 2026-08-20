import os
import shutil
import tempfile
import time
import unittest
from datetime import datetime
from zoneinfo import ZoneInfo

import openpyxl
import pandas as pd
import pytz

os.environ.setdefault("DATA_DIR", tempfile.mkdtemp(prefix="ct_dl_test_"))

import config
import data_layer as dl
import settings


HEADERS = ["DATE", "CUSTOMER NAME", "BANK NO.", "ACCOUNT NO.", "AMOUNT", "PAID TO"]


def build_workbook(name="t.xlsx", sheet="Sheet1", rows=None, template=None):
    path = config.DATA_DIR / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    for i, h in enumerate(HEADERS, 1):
        ws.cell(row=1, column=i, value=h)
    if template:
        for col, value in template.items():
            ws.cell(row=2, column=HEADERS.index(col) + 1, value=value)
    for excel_row, values in (rows or {}).items():
        for i, h in enumerate(HEADERS, 1):
            ws.cell(row=excel_row, column=i, value=values.get(h))
    wb.save(path)
    return path


class DataLayerTests(unittest.TestCase):
    def setUp(self):
        config.DATA_DIR.mkdir(parents=True, exist_ok=True)
        for old in list(config.DATA_DIR.glob("*")):
            if old.is_dir():
                shutil.rmtree(old)
            else:
                old.unlink()
        settings.set_append_direction("t.xlsx", "bottom")

    def test_row2_preserved_across_add(self):
        build_workbook(template={"AMOUNT": "=100*2", "PAID TO": "template note"})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        dl.add_row(sheet, {"DATE": "2026-08-02", "CUSTOMER NAME": "raju"})
        wb = openpyxl.load_workbook(config.DATA_DIR / "t.xlsx", data_only=False)
        ws = wb["Sheet1"]
        self.assertEqual(ws.cell(row=2, column=5).value, "=100*2")
        self.assertEqual(ws.cell(row=2, column=6).value, "template note")
        self.assertEqual(ws.cell(row=3, column=2).value, "raju")

    def test_append_bottom_and_top(self):
        build_workbook(rows={3: {"CUSTOMER NAME": "a"}})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        res = dl.add_row(sheet, {"CUSTOMER NAME": "b"})
        self.assertEqual(res["appended_to"], "bottom")
        self.assertEqual(res["excel_row"], 4)
        self.assertEqual(sheet.df.iloc[1]["CUSTOMER NAME"], "b")

        settings.set_append_direction("t.xlsx", "top")
        sheet2 = dl.load_sheet("t.xlsx", "Sheet1")
        res2 = dl.add_row(sheet2, {"CUSTOMER NAME": "c"})
        self.assertEqual(res2["appended_to"], "top")
        self.assertEqual(res2["excel_row"], 3)
        self.assertEqual(sheet2.df.iloc[0]["CUSTOMER NAME"], "c")

    def test_numeric_default_zero(self):
        build_workbook(rows={3: {"AMOUNT": "500"}})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        res = dl.add_row(sheet, {"DATE": "2026-08-02", "CUSTOMER NAME": "x", "AMOUNT": ""})
        self.assertEqual(sheet.df.iloc[-1]["AMOUNT"], "0")
        wb = openpyxl.load_workbook(config.DATA_DIR / "t.xlsx", data_only=False)
        self.assertEqual(wb["Sheet1"].cell(row=res["excel_row"], column=5).value, 0.0)

    def test_account_number_leading_zeros_preserved(self):
        build_workbook(rows={3: {"ACCOUNT NO.": "004578901234"}})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        res = dl.add_row(sheet, {"ACCOUNT NO.": "0099887766", "CUSTOMER NAME": "y"})
        wb = openpyxl.load_workbook(config.DATA_DIR / "t.xlsx", data_only=False)
        ws = wb["Sheet1"]
        self.assertEqual(ws.cell(row=3, column=4).value, "004578901234")
        self.assertEqual(ws.cell(row=res["excel_row"], column=4).value, "0099887766")

    def test_full_timestamp_roundtrip_as_datetime_cell(self):
        build_workbook()
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        res = dl.add_row(sheet, {"DATE": "2026-08-02 14:30:00", "CUSTOMER NAME": "new"})
        wb = openpyxl.load_workbook(config.DATA_DIR / "t.xlsx", data_only=False)
        cell = wb["Sheet1"].cell(row=res["excel_row"], column=1)
        self.assertIsInstance(cell.value, datetime)
        self.assertEqual(cell.value.strftime("%Y-%m-%d %H:%M:%S"), "2026-08-02 14:30:00")

    def test_old_text_date_stays_text(self):
        build_workbook(rows={3: {"DATE": "2026-08-01"}})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        dl.add_row(sheet, {"DATE": "2026-08-02 09:00:00", "CUSTOMER NAME": "z"})
        wb = openpyxl.load_workbook(config.DATA_DIR / "t.xlsx", data_only=False)
        ws = wb["Sheet1"]
        self.assertIsInstance(ws.cell(row=3, column=1).value, str)
        self.assertIsInstance(ws.cell(row=4, column=1).value, datetime)

    def test_duplicate_detection_and_force(self):
        build_workbook(rows={3: {"CUSTOMER NAME": "manu", "PHONE_X": "9"}})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        with self.assertRaises(dl.DuplicateError) as ctx:
            dl.add_row(sheet, {"CUSTOMER NAME": "MANU"})
        self.assertTrue(ctx.exception.duplicates)
        res = dl.add_row(sheet, {"CUSTOMER NAME": "MANU"}, force=True)
        self.assertEqual(len(res["duplicates"]), 1)
        self.assertEqual(sheet.row_count, 2)

    def test_duplicate_excludes_edited_row(self):
        build_workbook(rows={3: {"CUSTOMER NAME": "manu"}})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        res = dl.update_row(sheet, 3, {"CUSTOMER NAME": "manu"})
        self.assertEqual(res["duplicates"], [])

    def test_edit_delete_roundtrip(self):
        build_workbook(rows={3: {"CUSTOMER NAME": "a"}, 4: {"CUSTOMER NAME": "b"}})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        dl.update_row(sheet, 3, {"CUSTOMER NAME": "a2", "AMOUNT": "300"})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        self.assertEqual(sheet.cell_value("CUSTOMER NAME", 0), "a2")
        dl.delete_row(sheet, 3)
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        self.assertEqual(sheet.row_count, 1)
        self.assertEqual(sheet.cell_value("CUSTOMER NAME", 0), "b")

    def test_flag_adds_column_and_toast_flag(self):
        build_workbook(rows={3: {"CUSTOMER NAME": "a"}})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        self.assertFalse(sheet.has_flagged)
        res = dl.toggle_flag(sheet, 3)
        self.assertTrue(res["flagged"])
        self.assertTrue(res["flagged_column_added"])
        wb = openpyxl.load_workbook(config.DATA_DIR / "t.xlsx", data_only=False)
        self.assertEqual(wb["Sheet1"].cell(row=1, column=7).value, "FLAGGED")
        self.assertEqual(wb["Sheet1"].cell(row=3, column=7).value, "TRUE")

        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        res2 = dl.toggle_flag(sheet, 3)
        self.assertFalse(res2["flagged"])
        self.assertFalse(res2["flagged_column_added"])

    def test_backup_rotation_keeps_five(self):
        build_workbook(rows={3: {"CUSTOMER NAME": "a"}})
        for i in range(7):
            sheet = dl.load_sheet("t.xlsx", "Sheet1")
            dl.add_row(sheet, {"CUSTOMER NAME": f"u{i}"})
        backups = dl.list_backups("t.xlsx")
        self.assertEqual(len(backups), 5)
        backup_dir = config.BAK_DIR
        self.assertEqual(len(list(backup_dir.glob("t.xlsx.*.bak"))), 5)

    def test_revert_to_backup_restores_snapshot(self):
        build_workbook(rows={3: {"CUSTOMER NAME": "a"}})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        dl.add_row(sheet, {"CUSTOMER NAME": "b"})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        dl.add_row(sheet, {"CUSTOMER NAME": "c"})

        backups = dl.list_backups("t.xlsx")
        self.assertEqual(len(backups), 2)
        target = backups[1]["filename"]

        dl.revert_to_backup("t.xlsx", target)

        reloaded = dl.load_sheet("t.xlsx", "Sheet1")
        self.assertEqual(reloaded.row_count, 2)
        names = [reloaded.cell_value("CUSTOMER NAME", i) for i in range(reloaded.row_count)]
        self.assertEqual(names, ["a", "b"])

        after = dl.list_backups("t.xlsx")
        self.assertEqual(len(after), 3)
        self.assertNotEqual(after[0]["filename"], target)

        audit = config.AUDIT_LOG.read_text(encoding="utf-8")
        self.assertIn("REVERT", audit)
        self.assertIn(target, audit)

    def test_revert_rejects_backup_from_another_workbook(self):
        build_workbook(rows={3: {"CUSTOMER NAME": "a"}})
        other_name = "other.xlsx.20260101-000000-000000.bak"
        config.BAK_DIR.mkdir(parents=True, exist_ok=True)
        (config.BAK_DIR / other_name).write_bytes(b"x")
        with self.assertRaises(dl.SheetError):
            dl.revert_to_backup("t.xlsx", other_name)

    def test_timestamps_are_ist_not_host_clock(self):
        old_tz = os.environ.get("TZ")
        os.environ["TZ"] = "UTC"
        time.tzset()
        try:
            build_workbook()
            expected = datetime.now(ZoneInfo("Asia/Kolkata"))

            stamp = dl.local_now_str()
            self.assertTrue(stamp.endswith(" IST"))
            stamp_dt = datetime.strptime(stamp[:19], "%Y-%m-%d %H:%M:%S")
            self.assertLess(
                abs((stamp_dt - expected.replace(tzinfo=None)).total_seconds()), 10
            )

            sheet = dl.load_sheet("t.xlsx", "Sheet1")
            dl.add_row(sheet, {"CUSTOMER NAME": "tzcheck"})
            backup = dl.list_backups("t.xlsx")[0]
            created = datetime.strptime(backup["created_at"], "%Y-%m-%d %H:%M:%S")
            self.assertLess(
                abs((created - expected.replace(tzinfo=None)).total_seconds()), 10
            )
        finally:
            if old_tz is None:
                os.environ.pop("TZ", None)
            else:
                os.environ["TZ"] = old_tz
            time.tzset()

    def test_audit_log_lines_and_redaction(self):
        build_workbook(rows={3: {"CUSTOMER NAME": "a", "BANK NO.": "111", "ACCOUNT NO.": "222"}})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        dl.add_row(sheet, {"CUSTOMER NAME": "b", "BANK NO.": "333", "ACCOUNT NO.": "444"})
        dl.toggle_flag(sheet, 3)
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        dl.delete_row(sheet, 3)
        lines = config.AUDIT_LOG.read_text(encoding="utf-8").strip().splitlines()
        self.assertEqual(len(lines), 3)
        self.assertIn("ADD", lines[0])
        self.assertIn("BANK NO.=<redacted>", lines[0])
        self.assertNotIn("333", lines[0])
        self.assertIn("FLAG", lines[1])
        self.assertIn("DELETE", lines[2])

    def test_missing_sheet_default_columns(self):
        build_workbook()
        sheet = dl.load_sheet("t.xlsx", "NoSuchSheet")
        self.assertEqual(sheet.headers, ["DATE", "CUSTOMER NAME"])
        self.assertEqual(sheet.row_count, 0)

    def test_add_sheet(self):
        build_workbook()
        dl.add_sheet("t.xlsx", "Sheet2")
        self.assertIn("Sheet2", dl.list_sheets("t.xlsx"))
        with self.assertRaises(dl.SheetError):
            dl.add_sheet("t.xlsx", "Sheet2")

    def test_add_sheet_with_declared_columns(self):
        build_workbook()
        dl.add_sheet(
            "t.xlsx", "Sheet2",
            [{"name": "Amount", "type": "number"}, {"name": "Renewal", "type": "date"}],
        )
        wb = openpyxl.load_workbook(config.DATA_DIR / "t.xlsx", data_only=False)
        ws = wb["Sheet2"]
        self.assertEqual(ws.cell(row=1, column=1).value, "Amount")
        self.assertEqual(ws.cell(row=1, column=2).value, "Renewal")
        sheet = dl.load_sheet("t.xlsx", "Sheet2")
        self.assertEqual(sheet.headers, ["Amount", "Renewal"])
        self.assertIn("Amount", sheet.numeric_cols)
        self.assertIn("Renewal", sheet.date_cols)
        self.assertEqual(
            settings.get_column_types("t.xlsx", "Sheet2"),
            {"Amount": "number", "Renewal": "date"},
        )

    def test_add_sheet_default_columns_when_omitted(self):
        build_workbook()
        dl.add_sheet("t.xlsx", "Sheet2")
        wb = openpyxl.load_workbook(config.DATA_DIR / "t.xlsx", data_only=False)
        self.assertEqual(wb["Sheet2"].cell(row=1, column=1).value, "DATE")
        self.assertEqual(wb["Sheet2"].cell(row=1, column=2).value, "CUSTOMER NAME")

    def test_declared_types_classify_empty_sheet(self):
        build_workbook()
        dl.add_sheet(
            "t.xlsx", "Empty",
            [{"name": "Amount", "type": "number"}, {"name": "Renewal", "type": "date"}],
        )
        sheet = dl.load_sheet("t.xlsx", "Empty")
        self.assertEqual(sheet.row_count, 0)
        self.assertIn("Amount", sheet.numeric_cols)
        self.assertIn("Renewal", sheet.date_cols)

    def test_add_column_with_type(self):
        build_workbook(rows={3: {"CUSTOMER NAME": "a"}})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        dl.add_column(sheet, "Phone", "number")
        reloaded = dl.load_sheet("t.xlsx", "Sheet1")
        self.assertIn("Phone", reloaded.headers)
        self.assertIn("Phone", reloaded.numeric_cols)
        wb = openpyxl.load_workbook(config.DATA_DIR / "t.xlsx", data_only=False)
        self.assertEqual(wb["Sheet1"].cell(row=1, column=7).value, "Phone")
        self.assertIsNone(wb["Sheet1"].cell(row=3, column=7).value)

    def test_add_sheet_with_amount_column(self):
        build_workbook()
        dl.add_sheet(
            "t.xlsx", "Sheet2",
            [{"name": "Balance", "type": "amount"}, {"name": "Renewal", "type": "date"}],
        )
        sheet = dl.load_sheet("t.xlsx", "Sheet2")
        self.assertIn("Balance", sheet.amount_cols)
        self.assertIn("Balance", sheet.numeric_cols)
        self.assertIn("Renewal", sheet.date_cols)
        self.assertEqual(
            settings.get_column_types("t.xlsx", "Sheet2"),
            {"Balance": "amount", "Renewal": "date"},
        )

    def test_add_column_with_amount_type(self):
        build_workbook(rows={3: {"CUSTOMER NAME": "a"}})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        dl.add_column(sheet, "Balance", "amount")
        reloaded = dl.load_sheet("t.xlsx", "Sheet1")
        self.assertIn("Balance", reloaded.amount_cols)
        self.assertIn("Balance", reloaded.numeric_cols)
        self.assertEqual(settings.get_column_types("t.xlsx", "Sheet1").get("Balance"), "amount")

    def test_set_column_type_to_amount(self):
        build_workbook(rows={3: {"AMOUNT": "500"}})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        self.assertIn("AMOUNT", sheet.numeric_cols)
        dl.set_column_type(sheet, "AMOUNT", "amount")
        reloaded = dl.load_sheet("t.xlsx", "Sheet1")
        self.assertIn("AMOUNT", reloaded.amount_cols)
        self.assertIn("AMOUNT", reloaded.numeric_cols)
        self.assertEqual(settings.get_column_types("t.xlsx", "Sheet1").get("AMOUNT"), "amount")

    def test_set_column_type_to_text_clears_amount(self):
        build_workbook(rows={3: {"AMOUNT": "500"}})
        settings.set_column_types("t.xlsx", "Sheet1", {"AMOUNT": "amount"})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        self.assertIn("AMOUNT", sheet.amount_cols)
        dl.set_column_type(sheet, "AMOUNT", "text")
        reloaded = dl.load_sheet("t.xlsx", "Sheet1")
        self.assertNotIn("AMOUNT", reloaded.amount_cols)
        self.assertIn("AMOUNT", reloaded.numeric_cols)

    def test_set_column_type_rejects_bad_type_or_unknown_column(self):
        build_workbook()
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        with self.assertRaises(dl.SheetError):
            dl.set_column_type(sheet, "AMOUNT", "money")
        with self.assertRaises(dl.SheetError):
            dl.set_column_type(sheet, "NOPE", "number")

    def test_rename_amount_column_preserves_type(self):
        build_workbook(rows={3: {"AMOUNT": "500"}})
        settings.set_column_types("t.xlsx", "Sheet1", {"AMOUNT": "amount"})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        dl.rename_column(sheet, "AMOUNT", "AMOUNT TOTAL")
        reloaded = dl.load_sheet("t.xlsx", "Sheet1")
        self.assertIn("AMOUNT TOTAL", reloaded.amount_cols)
        self.assertNotIn("AMOUNT", reloaded.amount_cols)
        self.assertEqual(settings.get_column_types("t.xlsx", "Sheet1"), {"AMOUNT TOTAL": "amount"})

    def test_delete_amount_column_removes_type(self):
        build_workbook(rows={3: {"AMOUNT": "500"}})
        settings.set_column_types("t.xlsx", "Sheet1", {"AMOUNT": "amount"})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        dl.delete_column(sheet, "AMOUNT")
        reloaded = dl.load_sheet("t.xlsx", "Sheet1")
        self.assertNotIn("AMOUNT", reloaded.amount_cols)
        self.assertEqual(settings.get_column_types("t.xlsx", "Sheet1"), {})

    def test_delete_column_removes_data_and_shifts(self):
        build_workbook(rows={3: {"DATE": "2026-08-01", "CUSTOMER NAME": "a",
                                  "AMOUNT": "500", "PAID TO": "note"}})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        dl.delete_column(sheet, "AMOUNT")
        reloaded = dl.load_sheet("t.xlsx", "Sheet1")
        self.assertNotIn("AMOUNT", reloaded.headers)
        self.assertEqual(reloaded.cell_value("CUSTOMER NAME", 0), "a")
        self.assertEqual(reloaded.cell_value("PAID TO", 0), "note")
        wb = openpyxl.load_workbook(config.DATA_DIR / "t.xlsx", data_only=False)
        ws = wb["Sheet1"]
        self.assertEqual(ws.cell(row=3, column=5).value, "note")
        self.assertIsNone(ws.cell(row=3, column=4).value)

    def test_rename_column_updates_header_and_types(self):
        build_workbook(rows={3: {"CUSTOMER NAME": "a"}})
        settings.set_column_types("t.xlsx", "Sheet1", {"CUSTOMER NAME": "text"})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        dl.rename_column(sheet, "CUSTOMER NAME", "NAME")
        reloaded = dl.load_sheet("t.xlsx", "Sheet1")
        self.assertIn("NAME", reloaded.headers)
        self.assertNotIn("CUSTOMER NAME", reloaded.headers)
        self.assertEqual(reloaded.cell_value("NAME", 0), "a")
        self.assertEqual(settings.get_column_types("t.xlsx", "Sheet1"), {"NAME": "text"})
        wb = openpyxl.load_workbook(config.DATA_DIR / "t.xlsx", data_only=False)
        self.assertEqual(wb["Sheet1"].cell(row=1, column=2).value, "NAME")

    def test_path_traversal_rejected(self):
        build_workbook()
        for bad in ("../evil.bak", "/etc/passwd", "..\\x.bak", ".hidden.bak"):
            with self.assertRaises(dl.SheetError):
                dl.backup_path(bad)


class FormulaColDetectionTests(unittest.TestCase):
    def setUp(self):
        config.DATA_DIR.mkdir(parents=True, exist_ok=True)
        for old in list(config.DATA_DIR.glob("*")):
            if old.is_dir():
                shutil.rmtree(old)
            else:
                old.unlink()
        settings.set_append_direction("t.xlsx", "bottom")

    def test_formula_cols_detected(self):
        build_workbook(template={"AMOUNT": "=M2-N2"})
        sheet = dl.load_sheet("t.xlsx", "Sheet1")
        self.assertIn("AMOUNT", sheet.formula_cols)


if __name__ == "__main__":
    unittest.main(verbosity=2)