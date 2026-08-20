import os
import shutil
import tempfile
import unittest

import openpyxl

os.environ.setdefault("DATA_DIR", tempfile.mkdtemp(prefix="ct_fml_test_"))

import config
import data_layer as dl
import settings
from formula import validate_formula


# 15 columns -> M is column 13, N is column 14, O is column 15 (formula col)
HEADERS = [
    "H1", "H2", "H3", "H4", "H5", "H6", "H7", "H8", "H9", "H10", "H11", "H12",
    "M_VAL", "N_VAL", "FORMULA_OUT",
]


def build_workbook(formula="=M2-N2", data_rows=None):
    path = config.DATA_DIR / "f.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for i, h in enumerate(HEADERS, 1):
        ws.cell(row=1, column=i, value=h)
    ws.cell(row=2, column=15, value=formula)
    for excel_row, values in (data_rows or {}).items():
        for col_idx, value in values.items():
            ws.cell(row=excel_row, column=col_idx, value=value)
    wb.save(path)
    return path


def add_new(sheet, values):
    return dl.add_row(sheet, values)


class FormulaEngineTests(unittest.TestCase):
    def setUp(self):
        config.DATA_DIR.mkdir(parents=True, exist_ok=True)
        for old in list(config.DATA_DIR.glob("*")):
            if old.is_dir():
                shutil.rmtree(old)
            else:
                old.unlink()
        settings.set_append_direction("f.xlsx", "bottom")

    def test_tier2_basic_arithmetic(self):
        build_workbook()
        sheet = dl.load_sheet("f.xlsx", "Sheet1")
        res = add_new(sheet, {"M_VAL": "10", "N_VAL": "4"})
        row = sheet.df.iloc[-1]
        self.assertEqual(row["FORMULA_OUT"], "6")
        self.assertEqual(res["warnings"], [])

    def test_empty_numeric_inputs_default_to_zero(self):
        build_workbook(data_rows={3: {13: "10", 14: "4", 15: "6"}})
        sheet = dl.load_sheet("f.xlsx", "Sheet1")
        add_new(sheet, {"M_VAL": "", "N_VAL": ""})
        self.assertEqual(sheet.df.iloc[-1]["FORMULA_OUT"], "0")

    def test_tier1_round(self):
        build_workbook(formula="=ROUND(M2/N2, 2)")
        sheet = dl.load_sheet("f.xlsx", "Sheet1")
        add_new(sheet, {"M_VAL": "10", "N_VAL": "3"})
        self.assertEqual(sheet.df.iloc[-1]["FORMULA_OUT"], "3.33")

    def test_tier1_if(self):
        build_workbook(formula="=IF(M2>N2,M2,N2)")
        sheet = dl.load_sheet("f.xlsx", "Sheet1")
        add_new(sheet, {"M_VAL": "10", "N_VAL": "4"})
        self.assertEqual(sheet.df.iloc[-1]["FORMULA_OUT"], "10")

    def test_tier1_sum_range(self):
        build_workbook(formula="=SUM(M2:N2)")
        sheet = dl.load_sheet("f.xlsx", "Sheet1")
        add_new(sheet, {"M_VAL": "10", "N_VAL": "4"})
        self.assertEqual(sheet.df.iloc[-1]["FORMULA_OUT"], "14")

    def test_tier1_vlookup(self):
        build_workbook(
            formula="=VLOOKUP(M2,B3:C4,2,FALSE)",
            data_rows={3: {2: 1, 3: "one"}, 4: {2: 2, 3: "two"}},
        )
        sheet = dl.load_sheet("f.xlsx", "Sheet1")
        add_new(sheet, {"M_VAL": "2", "N_VAL": ""})
        self.assertEqual(sheet.df.iloc[-1]["FORMULA_OUT"], "two")

    def test_unsafe_formula_rejected_with_warning(self):
        build_workbook(formula="=M2+EVIL()")
        sheet = dl.load_sheet("f.xlsx", "Sheet1")
        res = add_new(sheet, {"M_VAL": "10", "N_VAL": "4"})
        self.assertEqual(sheet.df.iloc[-1]["FORMULA_OUT"], "")
        self.assertTrue(res["warnings"])

    def test_missing_column_ref_skip_with_warning(self):
        build_workbook(formula="=Z2-N2")
        sheet = dl.load_sheet("f.xlsx", "Sheet1")
        res = add_new(sheet, {"M_VAL": "10", "N_VAL": "4"})
        self.assertEqual(sheet.df.iloc[-1]["FORMULA_OUT"], "")
        self.assertTrue(res["warnings"])

    def test_formula_col_not_overwritten_on_edit(self):
        build_workbook()
        sheet = dl.load_sheet("f.xlsx", "Sheet1")
        add_new(sheet, {"M_VAL": "10", "N_VAL": "4"})
        sheet = dl.load_sheet("f.xlsx", "Sheet1")
        res = dl.update_row(sheet, 3, {"M_VAL": "20", "N_VAL": "5", "FORMULA_OUT": "999"})
        self.assertEqual(sheet.df.iloc[0]["FORMULA_OUT"], "15")
        self.assertEqual(res["warnings"], [])

    def test_tier1_top_insert(self):
        build_workbook(data_rows={3: {13: "100", 14: "50", 15: "50"}})
        settings.set_append_direction("f.xlsx", "top")
        sheet = dl.load_sheet("f.xlsx", "Sheet1")
        add_new(sheet, {"M_VAL": "10", "N_VAL": "4"})
        self.assertEqual(sheet.df.iloc[0]["FORMULA_OUT"], "6")
        self.assertEqual(sheet.df.iloc[1]["FORMULA_OUT"], "50")


class TierBasicTests(unittest.TestCase):
    def setUp(self):
        config.DATA_DIR.mkdir(parents=True, exist_ok=True)
        for old in list(config.DATA_DIR.glob("*")):
            if old.is_dir():
                shutil.rmtree(old)
            else:
                old.unlink()
        settings.set_append_direction("f.xlsx", "bottom")

    def _assert_computes(self, formula, expected, values=None):
        build_workbook(formula=formula)
        sheet = dl.load_sheet("f.xlsx", "Sheet1")
        res = add_new(sheet, values or {"M_VAL": "10", "N_VAL": "4"})
        self.assertEqual(sheet.df.iloc[-1]["FORMULA_OUT"], expected)
        self.assertEqual(res["warnings"], [])

    def test_sum_range(self):
        self._assert_computes("=SUM(M2:N2)", "14")

    def test_count_range(self):
        self._assert_computes("=COUNT(M2:N2)", "2")

    def test_max_min(self):
        self._assert_computes("=MAX(M2:N2)", "10")
        self._assert_computes("=MIN(M2:N2)", "4")

    def test_median_average(self):
        self._assert_computes("=MEDIAN(M2,N2)", "7")
        self._assert_computes("=AVERAGE(M2,N2)", "7")

    def test_counta_countblank(self):
        self._assert_computes("=COUNTA(M2:N2)", "2")
        self._assert_computes("=COUNTBLANK(M2,N2)", "0")
        self._assert_computes("=COUNTBLANK(M2,N2)", "2", values={"M_VAL": "", "N_VAL": ""})

    def test_sumif_countif(self):
        self._assert_computes("=SUMIF(M2:N2,\">5\")", "10")
        self._assert_computes("=COUNTIF(M2:N2,\">4\")", "1")

    def test_disabled_stat_function_blank_with_warning(self):
        build_workbook(formula="=NORM.DIST(M2,0,1,TRUE)")
        sheet = dl.load_sheet("f.xlsx", "Sheet1")
        res = add_new(sheet, {"M_VAL": "10", "N_VAL": "4"})
        self.assertEqual(sheet.df.iloc[-1]["FORMULA_OUT"], "")
        self.assertTrue(res["warnings"])

    def test_mixed_workbook_tier_basic_survives_disabled_column(self):
        path = config.DATA_DIR / "f.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for i, h in enumerate(HEADERS, 1):
            ws.cell(row=1, column=i, value=h)
        ws.cell(row=2, column=14, value="=SUM(L2:M2)")
        ws.cell(row=2, column=15, value="=NORM.DIST(M2,0,1,TRUE)")
        wb.save(path)
        sheet = dl.load_sheet("f.xlsx", "Sheet1")
        res = add_new(sheet, {"H12": "30", "M_VAL": "10", "N_VAL": "4"})
        self.assertEqual(sheet.df.iloc[-1]["N_VAL"], "40")
        self.assertEqual(sheet.df.iloc[-1]["FORMULA_OUT"], "")
        self.assertTrue(any("FORMULA_OUT" in w for w in res["warnings"]))
        self.assertFalse(any("N_VAL" in w for w in res["warnings"]))

    def test_date_function_survives_scipy_removal(self):
        build_workbook(formula="=YEAR(DATE(2024,1,15))")
        sheet = dl.load_sheet("f.xlsx", "Sheet1")
        add_new(sheet, {"M_VAL": "10", "N_VAL": "4"})
        self.assertEqual(sheet.df.iloc[-1]["FORMULA_OUT"], "2024")

    def test_validate_accepts_tier_basic(self):
        build_workbook(formula="=SUM(M2:N2)")
        sheet = dl.load_sheet("f.xlsx", "Sheet1")
        ok, _ = validate_formula(sheet, "FORMULA_OUT", "=COUNT(M2:N2)")
        self.assertTrue(ok)
        ok, _ = validate_formula(sheet, "FORMULA_OUT", "=SUMIF(M2:N2,\">5\")")
        self.assertTrue(ok)

    def test_validate_rejects_disabled_function(self):
        build_workbook(formula="=SUM(M2:N2)")
        sheet = dl.load_sheet("f.xlsx", "Sheet1")
        ok, _ = validate_formula(sheet, "FORMULA_OUT", "=NORM.DIST(M2,0,1,TRUE)")
        self.assertFalse(ok)


if __name__ == "__main__":
    unittest.main(verbosity=2)