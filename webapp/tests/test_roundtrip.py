import os
import shutil
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from werkzeug.security import generate_password_hash

os.environ.setdefault("SECRET_KEY", "test-secret-key")
os.environ.setdefault("PASSWORD_HASH", generate_password_hash("secret123"))
os.environ.setdefault("LOGIN_RATE_LIMIT_BURST", "1000 per hour")
os.environ.setdefault("LOGIN_RATE_LIMIT_SUSTAINED", "1000 per hour")
os.environ.setdefault("DATA_DIR", tempfile.mkdtemp(prefix="ct_rt_"))

import config
import data_layer
import settings

CUSTOMERS_HEADERS = [
    "DATE", "CUSTOMER NAME", "CONTACT", "CAR", "MODEL",
    "AMOUNT", "TOTAL", "BANK NO.", "ACCOUNT NO.",
]

ENQUIRY_HEADERS = ["DATE", "CUSTOMER NAME", "CONTACT NUMBER", "EMAIL", "STATUS"]


def build_customers(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for i, header in enumerate(CUSTOMERS_HEADERS, 1):
        ws.cell(row=1, column=i, value=header)
    ws.cell(row=2, column=7, value="=F2*2")
    data = [
        ("2024-06-15 14:30:00", "manu", "9812345678", "ALTO", "LXI", 5000, "123456789", "1111222233334444"),
        ("2024-07-01 09:15:00", "sita", "9822334455", "WAGONR", "VXI", 7500.5, "987654321", "9999888877776666"),
        ("2024-08-20 18:45:00", "ravi", "9876543210", "SWIFT", "ZXI", 12000, "555555555", "4444333322221111"),
    ]
    for r, (date, name, contact, car, model, amount, bank, account) in enumerate(data, start=3):
        ws.cell(row=r, column=1, value=datetime.strptime(date, "%Y-%m-%d %H:%M:%S"))
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=contact)
        ws.cell(row=r, column=4, value=car)
        ws.cell(row=r, column=5, value=model)
        ws.cell(row=r, column=6, value=amount)
        ws.cell(row=r, column=8, value=bank)
        ws.cell(row=r, column=9, value=account)
    wb.save(path)


def build_enquiry(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for i, header in enumerate(ENQUIRY_HEADERS, 1):
        ws.cell(row=1, column=i, value=header)
    ws.cell(row=2, column=5, value="NEW")
    data = [
        ("2024-09-01 10:00:00", "priya", "9810000001", "priya@example.com", "NEW"),
        ("2024-09-02 11:00:00", "arjun", "9810000002", "arjun@example.com", "PENDING"),
        ("2024-09-03 12:00:00", "kavita", "9810000003", "kavita@example.com", "DONE"),
    ]
    for r, (date, name, contact, email, status) in enumerate(data, start=3):
        ws.cell(row=r, column=1, value=datetime.strptime(date, "%Y-%m-%d %H:%M:%S"))
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=contact)
        ws.cell(row=r, column=4, value=email)
        ws.cell(row=r, column=5, value=status)
    wb.save(path)


def canon(value) -> tuple:
    if value is None or value == "" or (isinstance(value, float) and pd.isna(value)):
        return ("",)
    if isinstance(value, str):
        if data_layer._NUMERIC_RE.match(value):
            return ("n", repr(float(value.replace(",", "."))))
        return ("s", value)
    if isinstance(value, datetime):
        return ("dt", value.strftime("%Y-%m-%d %H:%M:%S"))
    return ("s", str(value))


def snapshot(sheet) -> tuple:
    headers = tuple(sheet.headers)
    template = tuple(
        (col, str(spec.get("value")), spec.get("data_type"))
        for col, spec in sorted(sheet.template.items())
    )
    rows = tuple(tuple(canon(v) for v in row) for row in sheet.df.values)
    classes = (tuple(sheet.formula_cols), tuple(sheet.numeric_cols), tuple(sheet.date_cols))
    return (headers, template, rows, classes, sheet.has_flagged)


class RoundTripTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.src_dir = Path(tempfile.mkdtemp(prefix="ct_rt_src_"))
        build_customers(cls.src_dir / "customers.xlsx")
        build_enquiry(cls.src_dir / "ENQUIRY LIST.xlsx")

    def setUp(self):
        config.DATA_DIR = Path(tempfile.mkdtemp(prefix="ct_rt_case_"))
        config.BAK_DIR = config.DATA_DIR / ".bak"
        config.AUDIT_LOG = config.DATA_DIR / "audit.log"
        settings._cache = None
        shutil.copy(self.src_dir / "customers.xlsx", config.DATA_DIR / "customers.xlsx")
        shutil.copy(self.src_dir / "ENQUIRY LIST.xlsx", config.DATA_DIR / "ENQUIRY LIST.xlsx")

    def audit_lines(self) -> list:
        if not config.AUDIT_LOG.exists():
            return []
        return config.AUDIT_LOG.read_text(encoding="utf-8").strip().splitlines()

    def test_customers_roundtrip_stable_without_ops(self):
        sheet = data_layer.load_sheet("customers.xlsx", "Sheet1")
        s0 = snapshot(sheet)
        data_layer.save_sheet(sheet)
        s1 = snapshot(data_layer.load_sheet("customers.xlsx", "Sheet1"))
        data_layer.save_sheet(data_layer.load_sheet("customers.xlsx", "Sheet1"))
        s2 = snapshot(data_layer.load_sheet("customers.xlsx", "Sheet1"))
        self.assertEqual(s0, s1)
        self.assertEqual(s1, s2)

    def test_enquiry_roundtrip_stable_without_ops(self):
        sheet = data_layer.load_sheet("ENQUIRY LIST.xlsx", "Sheet1")
        s0 = snapshot(sheet)
        data_layer.save_sheet(sheet)
        s1 = snapshot(data_layer.load_sheet("ENQUIRY LIST.xlsx", "Sheet1"))
        data_layer.save_sheet(data_layer.load_sheet("ENQUIRY LIST.xlsx", "Sheet1"))
        s2 = snapshot(data_layer.load_sheet("ENQUIRY LIST.xlsx", "Sheet1"))
        self.assertEqual(s0, s1)
        self.assertEqual(s1, s2)

    def test_operations_preserve_contract(self):
        sheet = data_layer.load_sheet("customers.xlsx", "Sheet1")
        self.assertEqual(sheet.row_count, 3)

        with self.assertRaises(data_layer.DuplicateError):
            data_layer.add_row(sheet, {"CUSTOMER NAME": "manu"})

        result = data_layer.add_row(sheet, {
            "DATE": "2024-09-10 08:00:00", "CUSTOMER NAME": "newbie", "CONTACT": "9811111111",
            "CAR": "CELERIO", "MODEL": "CORE", "AMOUNT": "9000",
            "BANK NO.": "111222333", "ACCOUNT NO.": "1234567890123456",
        })
        self.assertEqual(result["appended_to"], "bottom")
        self.assertEqual(result["excel_row"], 6)

        settings.set_append_direction("customers.xlsx", "top")
        sheet = data_layer.load_sheet("customers.xlsx", "Sheet1")
        result = data_layer.add_row(sheet, {
            "DATE": "2024-09-11 09:30:00", "CUSTOMER NAME": "first", "CONTACT": "9811111112",
            "CAR": "EECO", "MODEL": "PRIME", "AMOUNT": "4500",
            "BANK NO.": "444555666", "ACCOUNT NO.": "9988776655443322",
        })
        self.assertEqual(result["appended_to"], "top")
        self.assertEqual(result["excel_row"], 3)

        sheet = data_layer.load_sheet("customers.xlsx", "Sheet1")
        self.assertEqual(sheet.row_count, 5)
        data_layer.update_row(sheet, 3, {"CUSTOMER NAME": "first-edit", "AMOUNT": "4600"})

        sheet = data_layer.load_sheet("customers.xlsx", "Sheet1")
        flag = data_layer.toggle_flag(sheet, 3)
        self.assertTrue(flag["flagged"])
        self.assertTrue(flag["flagged_column_added"])

        sheet = data_layer.load_sheet("customers.xlsx", "Sheet1")
        unflag = data_layer.toggle_flag(sheet, 3)
        self.assertFalse(unflag["flagged"])

        sheet = data_layer.load_sheet("customers.xlsx", "Sheet1")
        data_layer.delete_row(sheet, 3)

        final = data_layer.load_sheet("customers.xlsx", "Sheet1")
        self.assertEqual(final.row_count, 4)
        self.assertTrue(final.has_flagged)
        self.assertEqual(final.headers[-1], config.RESERVED_COLUMN)
        self.assertEqual(final.cell_value("CUSTOMER NAME", 0), "manu")
        self.assertEqual(final.cell_value("CUSTOMER NAME", 3), "newbie")
        self.assertEqual(final.cell_value("DATE", 3), "2024-09-10 08:00:00")

        data_layer.save_sheet(final)
        s1 = snapshot(data_layer.load_sheet("customers.xlsx", "Sheet1"))
        data_layer.save_sheet(data_layer.load_sheet("customers.xlsx", "Sheet1"))
        s2 = snapshot(data_layer.load_sheet("customers.xlsx", "Sheet1"))
        self.assertEqual(s1, s2)

        lines = self.audit_lines()
        actions = [line.split(" | ")[1] for line in lines]
        self.assertEqual(actions, ["ADD", "ADD", "EDIT", "FLAG", "UNFLAG", "DELETE"])
        for line in lines:
            if line.split(" | ")[1] == "ADD":
                self.assertIn("BANK NO.=<redacted>", line)
                self.assertIn("ACCOUNT NO.=<redacted>", line)
                self.assertNotIn("111222333", line)
                self.assertNotIn("1234567890123456", line)

        backups = data_layer.list_backups("customers.xlsx")
        self.assertEqual(len(backups), config.BACKUP_KEEP)

    def test_enquiry_delete_and_audit(self):
        sheet = data_layer.load_sheet("ENQUIRY LIST.xlsx", "Sheet1")
        data_layer.delete_row(sheet, 4)

        final = data_layer.load_sheet("ENQUIRY LIST.xlsx", "Sheet1")
        self.assertEqual(final.row_count, 2)
        self.assertEqual(final.cell_value("CUSTOMER NAME", 0), "priya")
        self.assertEqual(final.cell_value("CUSTOMER NAME", 1), "kavita")

        data_layer.save_sheet(final)
        s1 = snapshot(data_layer.load_sheet("ENQUIRY LIST.xlsx", "Sheet1"))
        data_layer.save_sheet(data_layer.load_sheet("ENQUIRY LIST.xlsx", "Sheet1"))
        s2 = snapshot(data_layer.load_sheet("ENQUIRY LIST.xlsx", "Sheet1"))
        self.assertEqual(s1, s2)

        lines = self.audit_lines()
        self.assertEqual(len(lines), 1)
        self.assertEqual(lines[0].split(" | ")[1], "DELETE")
        self.assertIn("CUSTOMER NAME=arjun", lines[0])


if __name__ == "__main__":
    unittest.main()