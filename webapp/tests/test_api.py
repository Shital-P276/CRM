import io
import json
import os
import tempfile
import unittest
from pathlib import Path

import openpyxl
from werkzeug.security import generate_password_hash

os.environ["SECRET_KEY"] = "test-secret-key"
os.environ["PASSWORD_HASH"] = generate_password_hash("secret123")
os.environ["LOGIN_RATE_LIMIT_BURST"] = "1000 per hour"
os.environ["LOGIN_RATE_LIMIT_SUSTAINED"] = "1000 per hour"
os.environ["DEFAULT_RATE_LIMIT"] = "10000 per hour"
os.environ.setdefault("DATA_DIR", tempfile.mkdtemp(prefix="ct_api_test_"))

import config
import data_layer
import settings
from app import create_app


class ApiTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config["TESTING"] = True
        cls.client = cls.app.test_client()

    def setUp(self):
        config.DATA_DIR = Path(tempfile.mkdtemp(prefix="ct_api_case_"))
        config.BAK_DIR = config.DATA_DIR / ".bak"
        config.AUDIT_LOG = config.DATA_DIR / "audit.log"
        settings._cache = None
        self.build_workbook()
        token = self._fresh_token()
        response = self.client.post(
            "/api/login", json={"password": "secret123"}, headers={"X-CSRF-Token": token}
        )
        self.assertEqual(response.status_code, 200)

    def build_workbook(self, name="a.xlsx"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for i, header in enumerate(["DATE", "CUSTOMER NAME", "BANK NO.", "ACCOUNT NO.", "AMOUNT", "PAID TO"], 1):
            ws.cell(row=1, column=i, value=header)
        ws.cell(row=2, column=5, value="=0")
        ws.cell(row=3, column=2, value="manu")
        ws.cell(row=3, column=5, value=500)
        wb.save(config.DATA_DIR / name)

    def _fresh_token(self):
        return self.client.get("/api/session").get_json()["csrf_token"]

    def _csrf(self, client=None):
        client = client or self.client
        return client.get("/api/session").get_json()["csrf_token"]

    def _post(self, url, payload):
        return self.client.post(url, json=payload, headers={"X-CSRF-Token": self._csrf()})

    def test_unauthenticated_401(self):
        fresh = self.app.test_client()
        self.assertEqual(fresh.get("/api/workbooks").status_code, 401)
        self.assertEqual(fresh.get("/api/sheet-data?wb=a.xlsx&sheet=Sheet1").status_code, 401)
        self.assertEqual(fresh.post("/api/rows", json={}).status_code, 401)

    def test_index_serves_app(self):
        response = self.client.get("/")
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Customer Tracker", response.data)
        self.assertIn(b"/static/app.js", response.data)
        js = self.client.get("/static/app.js")
        self.assertEqual(js.status_code, 200)
        js.close()
        css = self.client.get("/static/style.css")
        self.assertEqual(css.status_code, 200)
        css.close()

    def test_sheets_list(self):
        response = self.client.get("/api/sheets?wb=a.xlsx")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["sheets"], ["Sheet1"])

    def test_wrong_password_401(self):
        fresh = self.app.test_client()
        token = fresh.get("/api/session").get_json()["csrf_token"]
        response = fresh.post("/api/login", json={"password": "nope"}, headers={"X-CSRF-Token": token})
        self.assertEqual(response.status_code, 401)

    def test_lockout_after_repeated_failures(self):
        original = config.LOGIN_RATE_LIMIT
        config.LOGIN_RATE_LIMIT = "3 per minute"
        env = {"REMOTE_ADDR": "203.0.113.99"}
        client = self.app.test_client()
        try:
            for _ in range(3):
                token = client.get("/api/session", environ_base=env).get_json()["csrf_token"]
                attempt = client.post(
                    "/api/login", json={"password": "nope"},
                    headers={"X-CSRF-Token": token}, environ_base=env,
                )
                self.assertEqual(attempt.status_code, 401)
            token = client.get("/api/session", environ_base=env).get_json()["csrf_token"]
            blocked = client.post(
                "/api/login", json={"password": "secret123"},
                headers={"X-CSRF-Token": token}, environ_base=env,
            )
            self.assertEqual(blocked.status_code, 429)
            body = blocked.get_json()
            self.assertIsInstance(body["retry_after"], int)
            self.assertGreater(body["retry_after"], 0)
            self.assertIn("Too many login attempts", body["error"])
            self.assertIn(str(body["retry_after"]), body["error"])
        finally:
            config.LOGIN_RATE_LIMIT = original

    def test_rate_limit_retry_after_reflects_window(self):
        original = config.LOGIN_RATE_LIMIT
        config.LOGIN_RATE_LIMIT = "3 per 1 minute"
        env = {"REMOTE_ADDR": "203.0.113.77"}
        client = self.app.test_client()
        try:
            for _ in range(3):
                token = client.get("/api/session", environ_base=env).get_json()["csrf_token"]
                attempt = client.post(
                    "/api/login", json={"password": "nope"},
                    headers={"X-CSRF-Token": token}, environ_base=env,
                )
                self.assertEqual(attempt.status_code, 401)
            token = client.get("/api/session", environ_base=env).get_json()["csrf_token"]
            blocked = client.post(
                "/api/login", json={"password": "nope"},
                headers={"X-CSRF-Token": token}, environ_base=env,
            )
            self.assertEqual(blocked.status_code, 429)
            retry_after = blocked.get_json()["retry_after"]
            self.assertGreater(retry_after, 40)
        finally:
            config.LOGIN_RATE_LIMIT = original

    def test_csrf_required_on_mutations(self):
        response = self.client.post("/api/rows", json={"wb": "a.xlsx", "sheet": "Sheet1", "values": {"CUSTOMER NAME": "x"}})
        self.assertEqual(response.status_code, 403)

    def test_add_row_and_sheet_data(self):
        response = self._post("/api/rows", {
            "wb": "a.xlsx", "sheet": "Sheet1",
            "values": {"DATE": "2026-08-02 10:30:00", "CUSTOMER NAME": "raju", "AMOUNT": "2500"},
        })
        self.assertEqual(response.status_code, 200)
        body = response.get_json()
        self.assertIn("excel_row", body)
        self.assertEqual(body["appended_to"], "bottom")

        data = self.client.get("/api/sheet-data?wb=a.xlsx&sheet=Sheet1").get_json()
        self.assertEqual(len(data["rows"]), 2)
        self.assertIn("raju", [r["values"]["CUSTOMER NAME"] for r in data["rows"]])
        self.assertIn("AMOUNT", data["numeric_cols"])
        self.assertIn("DATE", data["date_cols"])

    def test_duplicate_409_then_force(self):
        response = self._post("/api/rows", {
            "wb": "a.xlsx", "sheet": "Sheet1", "values": {"CUSTOMER NAME": "MANU"},
        })
        self.assertEqual(response.status_code, 409)
        self.assertTrue(response.get_json()["duplicates"])
        force = self._post("/api/rows", {
            "wb": "a.xlsx", "sheet": "Sheet1", "values": {"CUSTOMER NAME": "MANU"}, "force": True,
        })
        self.assertEqual(force.status_code, 200)

    def test_edit_delete_flag(self):
        added = self._post("/api/rows", {
            "wb": "a.xlsx", "sheet": "Sheet1", "values": {"CUSTOMER NAME": "b"},
        }).get_json()
        row = added["excel_row"]
        edited = self.client.put(
            f"/api/rows/{row}", json={"wb": "a.xlsx", "sheet": "Sheet1", "values": {"CUSTOMER NAME": "b2"}},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(edited.status_code, 200)
        flag = self._post(f"/api/rows/{row}/flag", {"wb": "a.xlsx", "sheet": "Sheet1"})
        self.assertEqual(flag.status_code, 200)
        self.assertTrue(flag.get_json()["flagged"])
        self.assertTrue(flag.get_json()["flagged_column_added"])
        deleted = self.client.delete(
            f"/api/rows/{row}", json={"wb": "a.xlsx", "sheet": "Sheet1"},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(deleted.status_code, 200)

    def test_traversal_rejected(self):
        response = self._post("/api/rows", {
            "wb": "../evil.xlsx", "sheet": "Sheet1", "values": {"CUSTOMER NAME": "x"},
        })
        self.assertEqual(response.status_code, 400)
        response = self._post("/api/rows", {
            "wb": "a.xlsx", "sheet": "Sheet1", "values": {"CUSTOMER NAME": "x"}, "force": True,
        })
        self.assertEqual(response.status_code, 200)

    def test_upload_valid_and_fake(self):
        fake = self.client.post(
            "/api/upload", data={"file": (io.BytesIO(b"not an xlsx"), "b.xlsx")},
            content_type="multipart/form-data", headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(fake.status_code, 400)

        wb = openpyxl.Workbook()
        wb.active.title = "Data"
        wb.active["A1"] = "X"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        good = self.client.post(
            "/api/upload", data={"file": (buf, "new.xlsx")},
            content_type="multipart/form-data", headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(good.status_code, 200)
        names = [w["name"] for w in good.get_json()["workbooks"]]
        self.assertIn("new.xlsx", names)

    def test_backups_list_and_download(self):
        self._post("/api/rows", {
            "wb": "a.xlsx", "sheet": "Sheet1", "values": {"CUSTOMER NAME": "c"},
        })
        backups = self.client.get("/api/backups?wb=a.xlsx").get_json()["backups"]
        self.assertEqual(len(backups), 1)
        filename = backups[0]["filename"]
        download = self.client.get(f"/api/backups/{filename}/download")
        self.assertEqual(download.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(download.data))
        wb.close()
        traversal = self.client.get("/api/backups/..%2Fevil.bak/download")
        self.assertIn(traversal.status_code, (400, 404))

    def test_backup_revert_api(self):
        self._post("/api/rows", {
            "wb": "a.xlsx", "sheet": "Sheet1", "values": {"CUSTOMER NAME": "a"},
        })
        self._post("/api/rows", {
            "wb": "a.xlsx", "sheet": "Sheet1", "values": {"CUSTOMER NAME": "b"},
        })
        backups = self.client.get("/api/backups?wb=a.xlsx").get_json()["backups"]
        self.assertEqual(len(backups), 2)
        target = backups[1]["filename"]

        revert = self.client.post(
            f"/api/backups/{target}/revert",
            json={"wb": "a.xlsx"},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(revert.status_code, 200)
        after = revert.get_json()["backups"]
        self.assertEqual(len(after), 3)

        data = self.client.get("/api/sheet-data?wb=a.xlsx&sheet=Sheet1").get_json()
        names = [r["values"]["CUSTOMER NAME"] for r in data["rows"]]
        self.assertEqual(names, ["manu", "a"])

        audit = config.AUDIT_LOG.read_text(encoding="utf-8")
        self.assertIn("REVERT", audit)

    def test_backup_revert_rejects_other_workbook(self):
        self._post("/api/rows", {
            "wb": "a.xlsx", "sheet": "Sheet1", "values": {"CUSTOMER NAME": "a"},
        })
        filename = self.client.get("/api/backups?wb=a.xlsx").get_json()["backups"][0]["filename"]
        revert = self.client.post(
            f"/api/backups/{filename}/revert",
            json={"wb": "other.xlsx"},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(revert.status_code, 400)

    def test_settings_toggle(self):
        put = self.client.put(
            "/api/settings", json={"wb": "a.xlsx", "append_direction": "top"},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(put.status_code, 200)
        self.assertEqual(put.get_json()["append_direction"], "top")
        get = self.client.get("/api/settings?wb=a.xlsx").get_json()
        self.assertEqual(get["append_direction"], "top")

    def test_settings_accepts_wb_query_param_like_sheet_data(self):
        """Frontend sends wb in the query string for PUT /api/settings (same
        style as /api/sheet-data); the route must not require it in the body."""
        sheet_data = self.client.get("/api/sheet-data?wb=a.xlsx&sheet=Sheet1")
        self.assertEqual(sheet_data.status_code, 200)
        get = self.client.get("/api/settings?wb=a.xlsx")
        self.assertEqual(get.status_code, 200)
        put = self.client.put(
            "/api/settings?wb=a.xlsx",
            json={"append_direction": "top"},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(put.status_code, 200)
        self.assertEqual(put.get_json()["append_direction"], "top")
        get2 = self.client.get("/api/settings?wb=a.xlsx").get_json()
        self.assertEqual(get2["append_direction"], "top")

    def test_duplicate_check_defaults_to_keyword_matching(self):
        dup_name = self._post("/api/rows", {
            "wb": "a.xlsx", "sheet": "Sheet1", "values": {"CUSTOMER NAME": "manu", "AMOUNT": "999"},
        })
        self.assertEqual(dup_name.status_code, 409)
        self.assertEqual(dup_name.get_json()["duplicates"][0]["column"], "CUSTOMER NAME")

        dup_amount = self._post("/api/rows", {
            "wb": "a.xlsx", "sheet": "Sheet1", "values": {"CUSTOMER NAME": "other", "AMOUNT": "500"},
        })
        self.assertEqual(dup_amount.status_code, 200)

    def test_duplicate_check_explicit_columns_override_keywords(self):
        put = self.client.put(
            "/api/settings?wb=a.xlsx&sheet=Sheet1",
            json={"append_direction": "top", "duplicate_check_columns": ["AMOUNT"]},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(put.status_code, 200)
        self.assertEqual(put.get_json()["duplicate_check_columns"], ["AMOUNT"])

        settings_data = json.loads((config.DATA_DIR / "settings.json").read_text())
        self.assertEqual(settings_data["a.xlsx"]["append_direction"], "top")
        self.assertEqual(
            settings_data["a.xlsx"]["duplicate_check_columns_by_sheet"]["Sheet1"], ["AMOUNT"]
        )

        ignored_keyword = self._post("/api/rows", {
            "wb": "a.xlsx", "sheet": "Sheet1",
            "values": {"CUSTOMER NAME": "manu", "AMOUNT": "999"},
        })
        self.assertEqual(ignored_keyword.status_code, 200)

        flagged = self._post("/api/rows", {
            "wb": "a.xlsx", "sheet": "Sheet1",
            "values": {"CUSTOMER NAME": "new", "AMOUNT": "500"},
        })
        self.assertEqual(flagged.status_code, 409)
        self.assertEqual(flagged.get_json()["duplicates"][0]["column"], "AMOUNT")

    def test_duplicate_check_rejects_unknown_column(self):
        put = self.client.put(
            "/api/settings?wb=a.xlsx&sheet=Sheet1",
            json={"append_direction": "bottom", "duplicate_check_columns": ["NOT A COLUMN"]},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(put.status_code, 400)
        self.assertIn("NOT A COLUMN", put.get_json()["error"])

    def test_get_settings_sheet_param_returns_config(self):
        response = self.client.get("/api/settings?wb=a.xlsx&sheet=Sheet1")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["duplicate_check_columns"], [])
        self.assertIn("append_direction", response.get_json())

    def test_last_opened_defaults_null(self):
        response = self.client.get("/api/last-opened")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json(), {"workbook": None, "sheet": None})

    def test_last_opened_roundtrip(self):
        put = self.client.put(
            "/api/last-opened",
            json={"workbook": "a.xlsx", "sheet": "Sheet1"},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(put.status_code, 200)
        self.assertEqual(put.get_json(), {"workbook": "a.xlsx", "sheet": "Sheet1"})

        get = self.client.get("/api/last-opened")
        self.assertEqual(get.status_code, 200)
        self.assertEqual(get.get_json(), {"workbook": "a.xlsx", "sheet": "Sheet1"})

        stored = json.loads((config.DATA_DIR / "settings.json").read_text())
        self.assertEqual(stored["last_opened"], {"workbook": "a.xlsx", "sheet": "Sheet1"})

    def test_last_opened_rejects_invalid_names(self):
        put = self.client.put(
            "/api/last-opened",
            json={"workbook": "../evil.xlsx", "sheet": "Sheet1"},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(put.status_code, 400)
        self.assertEqual(self.client.get("/api/last-opened").get_json(),
                         {"workbook": None, "sheet": None})

    def test_last_opened_stale_workbook_returns_null_not_error(self):
        put = self.client.put(
            "/api/last-opened",
            json={"workbook": "a.xlsx", "sheet": "Sheet1"},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(put.status_code, 200)
        (config.DATA_DIR / "a.xlsx").unlink()

        response = self.client.get("/api/last-opened")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json(), {"workbook": None, "sheet": None})

    def test_last_opened_stale_sheet_returns_null_sheet_not_error(self):
        put = self.client.put(
            "/api/last-opened",
            json={"workbook": "a.xlsx", "sheet": "Gone Sheet"},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(put.status_code, 200)

        response = self.client.get("/api/last-opened")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json(), {"workbook": "a.xlsx", "sheet": None})

    def test_last_opened_fresh_login_lands_on_remembered(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet2"
        for i, header in enumerate(["DATE", "CUSTOMER NAME"], 1):
            ws.cell(row=1, column=i, value=header)
        wb.save(config.DATA_DIR / "b.xlsx")

        put = self.client.put(
            "/api/last-opened",
            json={"workbook": "b.xlsx", "sheet": "Sheet2"},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(put.status_code, 200)

        fresh = self.app.test_client()
        token = fresh.get("/api/session").get_json()["csrf_token"]
        login = fresh.post(
            "/api/login", json={"password": "secret123"}, headers={"X-CSRF-Token": token}
        )
        self.assertEqual(login.status_code, 200)

        workbooks = [w["name"] for w in fresh.get("/api/workbooks").get_json()["workbooks"]]
        self.assertIn("b.xlsx", workbooks)
        remembered = fresh.get("/api/last-opened").get_json()
        self.assertEqual(remembered, {"workbook": "b.xlsx", "sheet": "Sheet2"})
        self.assertIn("Sheet2", fresh.get("/api/sheets?wb=b.xlsx").get_json()["sheets"])

    def test_security_headers_present(self):
        response = self.client.get("/api/session")
        self.assertEqual(response.headers["X-Frame-Options"], "DENY")
        self.assertEqual(response.headers["X-Content-Type-Options"], "nosniff")
        self.assertEqual(response.headers["Cache-Control"], "no-store")

    def test_create_sheet_with_typed_columns(self):
        post = self.client.post(
            "/api/sheets",
            json={"wb": "a.xlsx", "name": "Typed",
                  "columns": [{"name": "Amount", "type": "number"},
                              {"name": "Renewal", "type": "date"}]},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(post.status_code, 200)
        self.assertIn("Typed", post.get_json()["sheets"])

        data = self.client.get("/api/sheet-data?wb=a.xlsx&sheet=Typed").get_json()
        self.assertEqual(data["headers"], ["Amount", "Renewal"])
        self.assertEqual(data["rows"], [])
        self.assertIn("Amount", data["numeric_cols"])
        self.assertIn("Renewal", data["date_cols"])
        self.assertNotIn("Renewal", data["numeric_cols"])

    def test_create_sheet_without_columns_keeps_default(self):
        post = self.client.post(
            "/api/sheets", json={"wb": "a.xlsx", "name": "Plain"},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(post.status_code, 200)
        data = self.client.get("/api/sheet-data?wb=a.xlsx&sheet=Plain").get_json()
        self.assertEqual(data["headers"], ["DATE", "CUSTOMER NAME"])

    def test_create_sheet_rejects_bad_column_type(self):
        post = self.client.post(
            "/api/sheets",
            json={"wb": "a.xlsx", "name": "Bad",
                  "columns": [{"name": "X", "type": "money"}]},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(post.status_code, 400)
        self.assertNotIn("Bad", self.client.get("/api/sheets?wb=a.xlsx").get_json()["sheets"])

    def test_add_column_api(self):
        post = self.client.post(
            "/api/columns",
            json={"wb": "a.xlsx", "sheet": "Sheet1", "name": "Phone", "type": "number"},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(post.status_code, 200)
        self.assertIn("Phone", post.get_json()["headers"])
        self.assertIn("Phone", post.get_json()["numeric_cols"])
        data = self.client.get("/api/sheet-data?wb=a.xlsx&sheet=Sheet1").get_json()
        self.assertIn("Phone", data["headers"])

    def test_rename_column_api(self):
        put = self.client.put(
            "/api/columns",
            json={"wb": "a.xlsx", "sheet": "Sheet1", "name": "AMOUNT", "new_name": "AMOUNT TOTAL"},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(put.status_code, 200)
        self.assertIn("AMOUNT TOTAL", put.get_json()["headers"])
        self.assertNotIn("AMOUNT", put.get_json()["headers"])
        data = self.client.get("/api/sheet-data?wb=a.xlsx&sheet=Sheet1").get_json()
        self.assertIn("AMOUNT TOTAL", data["headers"])
        self.assertIn("AMOUNT TOTAL", data["formula_cols"])

    def test_delete_column_removes_data_and_backs_up(self):
        self._post("/api/rows", {
            "wb": "a.xlsx", "sheet": "Sheet1",
            "values": {"DATE": "2026-08-01", "CUSTOMER NAME": "keep"},
        })
        before = len(self.client.get("/api/backups?wb=a.xlsx").get_json()["backups"])

        delete = self.client.delete(
            "/api/columns", json={"wb": "a.xlsx", "sheet": "Sheet1", "name": "BANK NO."},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(delete.status_code, 200)
        self.assertNotIn("BANK NO.", delete.get_json()["headers"])

        data = self.client.get("/api/sheet-data?wb=a.xlsx&sheet=Sheet1").get_json()
        self.assertNotIn("BANK NO.", data["headers"])
        for row in data["rows"]:
            self.assertNotIn("BANK NO.", row["values"])

        backups = self.client.get("/api/backups?wb=a.xlsx").get_json()["backups"]
        self.assertEqual(len(backups), before + 1)

    def test_delete_column_rejects_flagged(self):
        delete = self.client.delete(
            "/api/columns", json={"wb": "a.xlsx", "sheet": "Sheet1", "name": "FLAGGED"},
            headers={"X-CSRF-Token": self._csrf()},
        )
        self.assertEqual(delete.status_code, 400)


if __name__ == "__main__":
    unittest.main(verbosity=2)