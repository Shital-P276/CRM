import os
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

import openpyxl
from werkzeug.security import generate_password_hash

os.environ["SECRET_KEY"] = "test-secret-key"
os.environ["PASSWORD_HASH"] = generate_password_hash("secret123")
os.environ["LOGIN_RATE_LIMIT_BURST"] = "1000 per hour"
os.environ["LOGIN_RATE_LIMIT_SUSTAINED"] = "1000 per hour"
os.environ.setdefault("DATA_DIR", tempfile.mkdtemp(prefix="ct_bugfix_"))

import config
import data_layer
import settings
from app import create_app


def build_standard_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for i, h in enumerate(["DATE", "CUSTOMER NAME", "AMOUNT", "GST TOTAL"], 1):
        ws.cell(row=1, column=i, value=h)
    ws.cell(row=2, column=4, value="=C2*1.18")
    ws.cell(row=3, column=1, value=datetime(2026, 8, 1))
    ws.cell(row=3, column=2, value="first")
    ws.cell(row=3, column=3, value=1000)
    ws.cell(row=3, column=4, value=1180.0)
    wb.save(path)


def build_titled_workbook(path: Path) -> None:
    """Real-world shape: title/blank rows above the table; header at row 14
    starting at column B; a merged title cell; data from row 15."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Copy of Pending"
    ws.merge_cells("B5:I5")
    ws["B5"] = "Debt Tracker"
    ws["B6"] = "01 May onwards"
    for i, h in enumerate(["Date", "Category", "Person", "Paid?", "Amount", "Description"], 1):
        ws.cell(row=14, column=1 + i, value=h)
    ws.cell(row=15, column=2, value=datetime(2025, 5, 1))
    ws.cell(row=15, column=3, value="To Me")
    ws.cell(row=15, column=4, value="Ritesh")
    ws.cell(row=15, column=5, value="Yes")
    ws.cell(row=15, column=6, value=1000)
    ws.cell(row=15, column=7, value="note")
    ws.cell(row=16, column=2, value=datetime(2025, 5, 11))
    ws.cell(row=16, column=3, value="From Me")
    ws.cell(row=16, column=4, value="Harsh")
    ws.cell(row=16, column=5, value="No")
    ws.cell(row=16, column=6, value=200)
    ws.cell(row=16, column=7, value="petrol")
    wb.save(path)


class BugFixTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config["TESTING"] = True
        cls.client = cls.app.test_client()

    def setUp(self):
        config.DATA_DIR = Path(tempfile.mkdtemp(prefix="ct_bugfix_case_"))
        config.BAK_DIR = config.DATA_DIR / ".bak"
        config.AUDIT_LOG = config.DATA_DIR / "audit.log"
        settings._cache = None

    def _login_headers(self):
        self.client.post("/api/login", json={"password": "secret123"})
        token = self.client.get("/api/session").get_json()["csrf_token"]
        return {"X-CSRF-Token": token}

    def test_titled_layout_reads_real_headers(self):
        build_titled_workbook(config.DATA_DIR / "Expenses .xlsx")
        sheet = data_layer.load_sheet("Expenses .xlsx", "Copy of Pending")
        self.assertEqual(
            sheet.headers,
            ["Date", "Category", "Person", "Paid?", "Amount", "Description"],
        )
        self.assertNotIn("Unnamed", " ".join(sheet.headers))
        self.assertEqual(sheet.header_row, 14)
        self.assertEqual(sheet.data_start_row, 15)
        self.assertEqual(sheet.first_col, 2)
        self.assertIn("Amount", sheet.numeric_cols)
        self.assertIn("Date", sheet.date_cols)

    def test_titled_layout_add_row_keeps_structure(self):
        build_titled_workbook(config.DATA_DIR / "Expenses .xlsx")
        settings.set_append_direction("Expenses .xlsx", "top")
        sheet = data_layer.load_sheet("Expenses .xlsx", "Copy of Pending")
        result = data_layer.add_row(
            sheet, {"Date": "2026-08-10", "Person": "NEW-TOP", "Amount": "999"}
        )
        self.assertEqual(result["excel_row"], 15)
        self.assertEqual(result["appended_to"], "top")

        wb = openpyxl.load_workbook(config.DATA_DIR / "Expenses .xlsx", data_only=False)
        ws = wb["Copy of Pending"]
        self.assertEqual(ws["B5"].value, "Debt Tracker")
        self.assertEqual(ws.cell(row=14, column=2).value, "Date")
        self.assertIsNone(ws.cell(row=14, column=1).value)
        self.assertEqual(ws.cell(row=15, column=4).value, "NEW-TOP")
        self.assertEqual(ws.cell(row=15, column=6).value, 999)

        reloaded = data_layer.load_sheet("Expenses .xlsx", "Copy of Pending")
        self.assertEqual(reloaded.headers[0], "Date")
        self.assertEqual(reloaded.row_count, 3)

    def test_append_direction_top_and_bottom(self):
        build_standard_workbook(config.DATA_DIR / "t.xlsx")
        headers = self._login_headers()

        self.client.put(
            "/api/settings", json={"wb": "t.xlsx", "append_direction": "top"}, headers=headers
        )
        r = self.client.post(
            "/api/rows",
            json={"wb": "t.xlsx", "sheet": "Sheet1",
                  "values": {"DATE": "2026-08-10", "CUSTOMER NAME": "top-row", "AMOUNT": "200"}},
            headers=headers,
        )
        data = r.get_json()
        self.assertEqual(data["excel_row"], 3)
        self.assertEqual(data["appended_to"], "top")

        self.client.put(
            "/api/settings", json={"wb": "t.xlsx", "append_direction": "bottom"}, headers=headers
        )
        r = self.client.post(
            "/api/rows",
            json={"wb": "t.xlsx", "sheet": "Sheet1",
                  "values": {"DATE": "2026-08-11", "CUSTOMER NAME": "bottom-row", "AMOUNT": "300"}},
            headers=headers,
        )
        data = r.get_json()
        self.assertEqual(data["excel_row"], 5)
        self.assertEqual(data["appended_to"], "bottom")

        wb = openpyxl.load_workbook(config.DATA_DIR / "t.xlsx", data_only=False)
        ws = wb["Sheet1"]
        self.assertEqual(ws.cell(row=3, column=2).value, "top-row")
        self.assertEqual(ws.cell(row=4, column=2).value, "first")
        self.assertEqual(ws.cell(row=5, column=2).value, "bottom-row")

    def test_formula_computed_end_to_end_api(self):
        build_standard_workbook(config.DATA_DIR / "t.xlsx")
        headers = self._login_headers()
        sheet = data_layer.load_sheet("t.xlsx", "Sheet1")
        self.assertIn("GST TOTAL", sheet.formula_cols)

        r = self.client.post(
            "/api/rows",
            json={"wb": "t.xlsx", "sheet": "Sheet1",
                  "values": {"DATE": "2026-08-12", "CUSTOMER NAME": "gst", "AMOUNT": "5000"}},
            headers=headers,
        )
        data = r.get_json()
        self.assertEqual(data["row"]["values"]["GST TOTAL"], "5900")
        self.assertEqual(data["warnings"], [])

        wb = openpyxl.load_workbook(config.DATA_DIR / "t.xlsx", data_only=False)
        ws = wb["Sheet1"]
        self.assertEqual(ws.cell(row=2, column=4).value, "=C2*1.18")
        self.assertEqual(ws.cell(row=4, column=4).value, 5900)


if __name__ == "__main__":
    unittest.main(verbosity=2)