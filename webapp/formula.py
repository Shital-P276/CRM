import ast
import copy
import io
import os
import re
import tempfile

import numpy as np
import openpyxl
import pandas as pd
import formulas
import tqdm
from openpyxl.cell.cell import MergedCell

_orig_tqdm_init = tqdm.tqdm.__init__


def _quiet_tqdm_init(self, *args, **kwargs):
    kwargs.setdefault("disable", True)
    return _orig_tqdm_init(self, *args, **kwargs)


tqdm.tqdm.__init__ = _quiet_tqdm_init

import sys as _sys
import types as _types
from formulas.functions import get_functions as _get_functions

# formulas 1.3.4 declares scipy as a base dependency, but scipy is only ever
# imported at module top level by functions/stat.py (scipy.stats/linalg/
# interpolate). This app uses neither statistical nor engineering functions,
# so stub those submodules out of sys.modules before formulas' get_functions()
# first imports them (comp.py is a legacy alias module built entirely on
# stat.py, so it is stubbed as well). scipy can then be removed entirely.
_stat_stub = _types.ModuleType("formulas.functions.stat")
_stat_stub.FUNCTIONS = {}  # type: ignore[attr-defined]
_eng_stub = _types.ModuleType("formulas.functions.eng")
_eng_stub.FUNCTIONS = {}  # type: ignore[attr-defined]
_comp_stub = _types.ModuleType("formulas.functions.comp")
_comp_stub.FUNCTIONS = {}  # type: ignore[attr-defined]
_sys.modules["formulas.functions.stat"] = _stat_stub
_sys.modules["formulas.functions.eng"] = _eng_stub
_sys.modules["formulas.functions.comp"] = _comp_stub
_get_functions.cache_clear()
del _sys, _types, _stat_stub, _eng_stub, _comp_stub, _get_functions

from data_layer import _NUMERIC_RE, _parse_datetime

_REF_RE = re.compile(r"\$?([A-Za-z]{1,3})\$?(\d+)")
_SAFE_RE = re.compile(r"^[0-9+\-*/().\s]+$")

_BIN_OPS = (ast.Add, ast.Sub, ast.Mult, ast.Div, ast.Pow, ast.Mod)
_UNARY_OPS = (ast.UAdd, ast.USub)
_ALLOWED_NODES = (
    ast.Expression, ast.BinOp, ast.UnaryOp, ast.Constant, ast.Load,
) + _BIN_OPS + _UNARY_OPS


def compute_formula_values(sheet, row_values, insert_mode, new_excel_row):
    formula_cols = [h for h in sheet.formula_cols]
    if not formula_cols:
        return {}, []

    translated = {}
    for header in formula_cols:
        col_idx = sheet.header_index(header)
        template = sheet.template.get(sheet.first_col + col_idx, {}).get("value")
        if isinstance(template, str) and template.startswith("="):
            translated[header] = _translate_row_refs(sheet, template, new_excel_row)

    if not translated:
        return {}, []

    result = {}
    warnings = []

    valid = {}
    for header, formula in translated.items():
        if _has_invalid_refs(formula, sheet):
            warnings.append(f"{header}: formula references a missing column")
        else:
            valid[header] = formula

    tier1 = _tier1(sheet, row_values, valid, formula_cols, insert_mode, new_excel_row)
    for header in formula_cols:
        if header not in valid:
            continue
        if tier1 is not None and tier1.get(header, ""):
            result[header] = tier1[header]
            continue
        value = _tier_basic(sheet, valid[header], row_values, new_excel_row, insert_mode)
        if value is not None:
            result[header] = value
            continue
        value = _tier2(sheet, row_values, valid[header], new_excel_row, insert_mode)
        if value is not None:
            result[header] = value
        else:
            warnings.append(f"{header}: formula could not be calculated")

    return result, warnings


def _sample_values(sheet) -> dict:
    sample = {h: "" for h in sheet.headers}
    if len(sheet.df) == 0:
        return sample
    last = sheet.df.iloc[-1]
    for header in sheet.headers:
        if header not in last.index:
            continue
        raw = last[header]
        sample[header] = "" if pd.isna(raw) else str(raw)
    return sample


def _probe_with_column(sheet, header):
    """Return (probe_sheet, tmp_path) for validating a formula whose column
    does not yet exist in the on-disk workbook. The probe writes the extra
    header into a throwaway copy so Tier 1 can evaluate it."""
    probe = copy.copy(sheet)
    probe.headers = list(sheet.headers) + [header]
    probe.template = dict(sheet.template)

    wb = openpyxl.load_workbook(sheet.path, data_only=False)
    ws = wb[sheet.sheet_name]
    col_idx = probe.header_index(header)
    ws.cell(row=probe.header_row, column=probe.first_col + col_idx, value=header)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    probe.path = tmp_path
    return probe, tmp_path


def validate_formula(sheet, header, formula_text) -> tuple:
    """Return (ok, reason). A formula is acceptable when it can be parsed and
    evaluated by the same safety path used for computing values — Tier 1
    (formulas lib) or Tier 2 (AST whitelist)."""
    formula = str(formula_text).strip()
    if not formula.startswith("="):
        return False, "formula must start with '='"

    probe = sheet
    tmp_path = None
    if header not in sheet.headers:
        probe, tmp_path = _probe_with_column(sheet, header)

    try:
        translated = _translate_row_refs(probe, formula, probe.data_start_row)
        if _has_invalid_refs(translated, probe):
            return False, "formula references a missing column"

        sample = _sample_values(probe)
        tier1 = _tier1(probe, sample, {header: translated}, [header], "edit", probe.data_start_row)
        if tier1 is not None and tier1.get(header) not in (None, ""):
            return True, ""
        value = _tier_basic(probe, translated, sample, probe.data_start_row, "edit")
        if value not in (None, ""):
            return True, ""
        value = _tier2(probe, sample, translated, probe.data_start_row, "edit")
        if value not in (None, ""):
            return True, ""
        return False, "formula could not be parsed or evaluated"
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass


def _has_invalid_refs(formula: str, sheet) -> bool:
    for match in _REF_RE.finditer(formula):
        col_1based = openpyxl.utils.column_index_from_string(match.group(1))  # type: ignore[attr-defined]
        if col_1based < sheet.first_col or col_1based > sheet.first_col + len(sheet.headers) - 1:
            return True
    return False


def _translate_row_refs(sheet, formula: str, new_excel_row: int) -> str:
    def repl(match):
        column, row = match.group(1), int(match.group(2))
        if row == sheet.template_row:
            return f"{column}{new_excel_row}"
        return f"{column}{row}"

    return _REF_RE.sub(repl, formula)


def _fmt(value) -> str:
    if isinstance(value, np.ndarray):
        if value.size == 0:
            return ""
        value = value.flat[0]
    if isinstance(value, np.bool_):
        value = bool(value)
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, np.integer):
        value = int(value)
    if isinstance(value, np.floating):
        value = float(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value != value or value in (float("inf"), float("-inf")):
            return ""
        if value.is_integer():
            return str(int(value))
        return str(round(value, 6))
    if value is None:
        return ""
    text = str(value)
    if text.startswith("#"):
        return ""
    return text


def _tier1(sheet, row_values, translated, formula_cols, insert_mode, new_excel_row):
    try:
        wb = openpyxl.load_workbook(sheet.path, data_only=False)
        ws = wb[sheet.sheet_name]
        if insert_mode == "top":
            ws.insert_rows(new_excel_row, 1)
        elif insert_mode == "edit":
            for col in range(1, max(ws.max_column, len(sheet.headers)) + 1):
                cell = ws.cell(row=new_excel_row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None

        for col_idx, header in enumerate(sheet.headers):
            cell = ws.cell(row=new_excel_row, column=sheet.first_col + col_idx)
            if isinstance(cell, MergedCell):
                continue
            if header in translated:
                cell.value = translated[header]
            else:
                raw = row_values.get(header, "")
                if raw == "" or raw is None:
                    cell.value = None
                elif header in sheet.date_cols:
                    parsed = _parse_datetime(str(raw))
                    cell.value = parsed if parsed else str(raw)
                elif _NUMERIC_RE.match(str(raw)):
                    cell.value = float(str(raw).replace(",", "."))
                else:
                    cell.value = str(raw)

        buffer = io.BytesIO()
        wb.save(buffer)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(buffer.getvalue())
            tmp_path = tmp.name

        try:
            model = formulas.ExcelModel().loads(tmp_path).finish()
            bookname = os.path.basename(tmp_path).lower()
            sheet_upper = sheet.sheet_name.upper()
            result = {}
            for header in formula_cols:
                try:
                    letter = openpyxl.utils.get_column_letter(sheet.first_col + sheet.header_index(header))  # type: ignore[attr-defined]
                    cell_id = f"'[{bookname}]{sheet_upper}'!{letter}{new_excel_row}"
                    cell = model.calculate(outputs=[cell_id]).get(cell_id)  # type: ignore[attr-defined]
                    result[header] = _fmt(cell.value) if cell is not None else ""
                except Exception:
                    result[header] = None
            return result
        finally:
            os.unlink(tmp_path)
    except Exception:
        return None


def _tier2(sheet, row_values, translated_formula, new_excel_row, insert_mode):
    if not translated_formula.startswith("="):
        return None
    try:
        expression = _resolve_refs(
            translated_formula[1:], sheet, row_values, new_excel_row, insert_mode
        )
        return _safe_arith(expression)
    except Exception:
        return None


def _resolve_refs(expression: str, sheet, row_values, new_excel_row, insert_mode) -> str:
    def repl(match):
        letter, row = match.group(1), int(match.group(2))
        col_1based = openpyxl.utils.column_index_from_string(letter)  # type: ignore[attr-defined]
        df_idx = col_1based - sheet.first_col
        if df_idx < 0 or df_idx >= len(sheet.headers):
            raise ValueError(f"invalid column reference: {letter}")
        header = sheet.headers[df_idx]
        if row == new_excel_row:
            value = row_values.get(header, "")
        else:
            value = _other_row_value(sheet, row, df_idx, insert_mode)
        if value in ("", None):
            value = _last_df_value(sheet, df_idx)
        if value in ("", None):
            return "0"
        return str(value)

    return _REF_RE.sub(repl, expression)


def _other_row_value(sheet, row: int, col_idx: int, insert_mode: str):
    if insert_mode == "top":
        df_idx = row - sheet.data_start_row - 1
    else:
        df_idx = row - sheet.data_start_row
    if df_idx < 0 or df_idx >= len(sheet.df):
        return ""
    raw = sheet.df.iat[df_idx, col_idx]
    return "" if pd.isna(raw) else str(raw)


def _last_df_value(sheet, col_idx: int):
    if len(sheet.df) == 0:
        return ""
    raw = sheet.df.iat[-1, col_idx]
    return "" if pd.isna(raw) else str(raw)


_FUNC_RE = re.compile(r"^([A-Z][A-Z0-9._]*)\s*\((.*)\)\s*$", re.DOTALL)

_TIER_BASIC_FUNCS = {
    "SUM", "SUMIF", "COUNT", "COUNTA", "COUNTBLANK", "COUNTIF",
    "AVERAGE", "AVERAGEA", "MEDIAN", "MAX", "MIN",
}

_RANGE_RE = re.compile(r"^\$?([A-Za-z]{1,3})\$?(\d+):\$?([A-Za-z]{1,3})\$?(\d+)$")


def _split_args(text: str) -> list:
    args, cur, depth, in_quotes = [], [], 0, False
    for ch in text:
        if ch == '"':
            in_quotes = not in_quotes
            cur.append(ch)
        elif ch == "(" and not in_quotes:
            depth += 1
            cur.append(ch)
        elif ch == ")" and not in_quotes:
            depth -= 1
            cur.append(ch)
        elif ch == "," and not in_quotes and depth == 0:
            args.append("".join(cur).strip())
            cur = []
        else:
            cur.append(ch)
    if cur:
        args.append("".join(cur).strip())
    return args


def _as_number(value):
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return None
    return None


def _cell_value(sheet, row_values, row, col_idx, insert_mode, new_excel_row, fallback=True):
    if col_idx < 0 or col_idx >= len(sheet.headers):
        return None
    header = sheet.headers[col_idx]
    if row == new_excel_row:
        value = row_values.get(header, "")
    else:
        value = _other_row_value(sheet, row, col_idx, insert_mode)
    if fallback and value in ("", None):
        value = _last_df_value(sheet, col_idx)
    return value


def _resolve_arg(sheet, row_values, arg, insert_mode, new_excel_row, fallback=True) -> list:
    if ":" in arg:
        match = _RANGE_RE.match(arg)
        if not match:
            return []
        start_col = _col_to_df_idx(match.group(1), sheet)
        end_col = _col_to_df_idx(match.group(3), sheet)
        start_row, end_row = int(match.group(2)), int(match.group(4))
        if start_col is None or end_col is None:
            return []
        if start_col > end_col:
            start_col, end_col = end_col, start_col
        if start_row > end_row:
            start_row, end_row = end_row, start_row
        values = []
        for row in range(start_row, end_row + 1):
            for col_idx in range(start_col, end_col + 1):
                value = _cell_value(sheet, row_values, row, col_idx, insert_mode, new_excel_row, fallback)
                if value is not None:
                    values.append(value)
        return values
    match = _REF_RE.match(arg)
    if match:
        col_idx = _col_to_df_idx(match.group(1), sheet)
        if col_idx is None:
            return []
        value = _cell_value(sheet, row_values, int(match.group(2)), col_idx, insert_mode, new_excel_row, fallback)
        return [value] if value is not None else []
    number = _as_number(arg)
    if number is not None:
        return [number]
    if arg.startswith('"') and arg.endswith('"') and len(arg) >= 2:
        return [arg[1:-1]]
    return []


def _col_to_df_idx(column: str, sheet):
    try:
        col_1based = openpyxl.utils.column_index_from_string(column)  # type: ignore[attr-defined]
    except ValueError:
        return None
    df_idx = col_1based - sheet.first_col
    if df_idx < 0 or df_idx >= len(sheet.headers):
        return None
    return df_idx


def _parse_criteria(text) -> tuple:
    text = str(text).strip()
    number = r"-?\d+(?:[.,]\d+)?"
    match = re.match(rf"^(>=|<=|<>|=|>|<)\s*({number})$", text)
    if match:
        return match.group(1), float(match.group(2).replace(",", "."))
    if re.match(rf"^{number}$", text):
        return "=", float(text.replace(",", "."))
    return "=", text


def _criteria_text(arg, sheet, row_values, insert_mode, new_excel_row):
    arg = arg.strip()
    if arg.startswith('"') and arg.endswith('"') and len(arg) >= 2:
        return arg[1:-1]
    resolved = _resolve_arg(sheet, row_values, arg, insert_mode, new_excel_row, fallback=False)
    if resolved:
        return resolved[0]
    return arg


def _match_criteria(value, criteria) -> bool:
    op, target = criteria
    number = _as_number(value)
    if isinstance(target, float):
        if number is None:
            return False
        if op == ">":
            return number > target
        if op == "<":
            return number < target
        if op == ">=":
            return number >= target
        if op == "<=":
            return number <= target
        if op == "<>":
            return number != target
        return number == target
    text = "" if value is None else str(value).strip()
    if op == "<>":
        return text != target
    return text == target


def _apply_basic(name, raw_args, sheet, row_values, insert_mode, new_excel_row):
    def values_of(arg):
        return _resolve_arg(sheet, row_values, arg, insert_mode, new_excel_row, fallback=False)

    if name in ("SUM", "COUNT", "COUNTA", "COUNTBLANK", "AVERAGE", "AVERAGEA", "MEDIAN", "MAX", "MIN"):
        flat = []
        for arg in raw_args:
            flat.extend(values_of(arg))
        numbers = [n for n in (_as_number(v) for v in flat) if n is not None]
        if name == "SUM":
            return sum(numbers)
        if name == "COUNT":
            return len(numbers)
        if name == "COUNTA":
            return sum(1 for v in flat if v not in ("", None))
        if name == "COUNTBLANK":
            return sum(1 for v in flat if v in ("", None))
        if name == "MAX":
            return max(numbers) if numbers else None
        if name == "MIN":
            return min(numbers) if numbers else None
        if name == "MEDIAN":
            if not numbers:
                return None
            numbers.sort()
            mid = len(numbers) // 2
            if len(numbers) % 2:
                return numbers[mid]
            return (numbers[mid - 1] + numbers[mid]) / 2.0
        if name == "AVERAGE":
            return sum(numbers) / len(numbers) if numbers else None
        if name == "AVERAGEA":
            nonempty = [v for v in flat if v not in ("", None)]
            if not nonempty:
                return None
            return sum(_as_number(v) or 0 for v in nonempty) / len(nonempty)

    if name == "COUNTIF":
        if len(raw_args) < 2:
            return None
        values = values_of(raw_args[0])
        criteria = _parse_criteria(_criteria_text(raw_args[1], sheet, row_values, insert_mode, new_excel_row))
        return sum(1 for v in values if _match_criteria(v, criteria))

    if name == "SUMIF":
        if len(raw_args) < 2:
            return None
        values = values_of(raw_args[0])
        criteria = _parse_criteria(_criteria_text(raw_args[1], sheet, row_values, insert_mode, new_excel_row))
        if len(raw_args) >= 3:
            sums = values_of(raw_args[2])
            total = 0.0
            for value, sum_value in zip(values, sums):
                if _match_criteria(value, criteria):
                    n = _as_number(sum_value)
                    if n is not None:
                        total += n
            return total
        total = 0.0
        for value in values:
            if _match_criteria(value, criteria):
                n = _as_number(value)
                if n is not None:
                    total += n
        return total

    return None


def _tier_basic(sheet, translated_formula, row_values, new_excel_row, insert_mode):
    if not translated_formula.startswith("="):
        return None
    match = _FUNC_RE.match(translated_formula[1:].strip())
    if not match:
        return None
    name = match.group(1).upper()
    if name not in _TIER_BASIC_FUNCS:
        return None
    args_text = match.group(2).strip()
    if not args_text:
        return None
    try:
        result = _apply_basic(name, _split_args(args_text), sheet, row_values, insert_mode, new_excel_row)
    except Exception:
        return None
    if result is None:
        return None
    return _fmt(result)


def _safe_arith(expression: str):
    if not _SAFE_RE.match(expression):
        raise ValueError("unsafe expression")
    tree = ast.parse(expression, mode="eval")
    for node in ast.walk(tree):
        if not isinstance(node, _ALLOWED_NODES):
            raise ValueError(f"unsupported node: {type(node).__name__}")
        if isinstance(node, ast.Constant) and not isinstance(node.value, (int, float)):
            raise ValueError("unsupported constant")
    result = eval(compile(tree, "<safe>", "eval"), {"__builtins__": {}}, {})
    if isinstance(result, complex):
        raise ValueError("complex result")
    return _fmt(result)