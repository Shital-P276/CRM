import re
import threading
import zipfile
from datetime import datetime
from zoneinfo import ZoneInfo
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.cell.cell import MergedCell

import config
import settings

EXCEL_HEADER_ROW = 1
TEMPLATE_ROW = 2
DATA_START_ROW = 3

_NUMERIC_RE = re.compile(r"^\s*-?\d+(?:[.,]\d+)?\s*$")
_DATETIME_RE = re.compile(
    r"^\s*(\d{4})-(\d{1,2})-(\d{1,2})(?:[ T](\d{1,2}):(\d{1,2})(?::(\d{1,2}))?)?\s*$"
)

DUP_KEYWORDS = (
    "CUSTOMER NAME", "CONTACT NUMBER", "CONTACT", "PHONE",
    "CAR", "VEHICLE", "MODEL", "EMAIL", "ACCOUNT",
)

IDENTIFIER_KEYWORDS = ("BANK", "ACCOUNT", "PHONE", "CONTACT")

DATE_NUMBER_FORMAT = "yyyy-mm-dd hh:mm:ss"

_default_columns = ["DATE", "CUSTOMER NAME"]

NEW_TYPE = "NEW"


class SheetError(Exception):
    pass


class DuplicateError(Exception):
    def __init__(self, duplicates):
        super().__init__("duplicates found")
        self.duplicates = duplicates


class FormulaValidationError(Exception):
    def __init__(self, errors):
        super().__init__("invalid formula")
        self.errors = errors


def local_now() -> datetime:
    return datetime.now(ZoneInfo(config.TIMEZONE))


def local_now_str() -> str:
    return local_now().strftime("%Y-%m-%d %H:%M:%S") + " IST"


def _audit(action: str, workbook: str, sheet: str, excel_row, description: str) -> None:
    config.DATA_DIR.mkdir(parents=True, exist_ok=True)
    line = f"{local_now_str()} | {action} | {workbook} | {sheet} | {excel_row} | {description}\n"
    with open(config.AUDIT_LOG, "a", encoding="utf-8") as fh:
        fh.write(line)
    _rotate_audit()


def _rotate_audit() -> None:
    path = config.AUDIT_LOG
    try:
        if path.exists() and path.stat().st_size > config.AUDIT_MAX_BYTES:
            backup = path.with_name("audit.log.1")
            if backup.exists():
                backup.unlink()
            path.rename(backup)
    except OSError:
        pass


def _redacted(values: dict, headers: list) -> str:
    parts = []
    for header in headers:
        value = values.get(header, "")
        if value == "":
            continue
        lowered = header.upper()
        if any(k in lowered for k in config.BANK_ACCOUNT_KEYWORDS):
            parts.append(f"{header}=<redacted>")
        else:
            parts.append(f"{header}={value}")
    return ", ".join(parts)


class Sheet:
    def __init__(self, workbook: str, sheet_name: str, path: Path, wb_obj, ws):
        self.workbook = workbook
        self.sheet_name = sheet_name
        self.path = path
        self.wb_obj = wb_obj
        self.ws = ws
        self.headers: list = []
        self.template: dict = {}
        self.df = pd.DataFrame()
        self.orig_types: list = []
        self.formula_cols: list = []
        self.numeric_cols: list = []
        self.date_cols: list = []
        self.has_flagged: bool = False
        self.warnings: list = []
        self.header_row: int = EXCEL_HEADER_ROW
        self.template_row: int = TEMPLATE_ROW
        self.data_start_row: int = DATA_START_ROW
        self.first_col: int = 1

    @property
    def row_count(self) -> int:
        return len(self.df)

    def excel_row_of(self, index: int) -> int:
        return self.data_start_row + index

    def header_index(self, header: str) -> int:
        return self.headers.index(header)

    def cell_value(self, header: str, index: int):
        value = self.df.iat[index, self.header_index(header)]
        return "" if pd.isna(value) else str(value)

    def set_cell(self, header: str, index: int, value) -> None:
        self.df.iat[index, self.header_index(header)] = value


def _capture_template(ws, template_row) -> dict:
    template = {}
    for cell in ws[template_row]:
        if isinstance(cell, MergedCell):
            continue
        template[cell.column] = {
            "value": cell.value,
            "data_type": cell.data_type,
            "number_format": cell.number_format,
        }
    return template


def _detect_layout(ws) -> tuple:
    """Find the real header row, whether the row below it is a template row,
    and the first table column.

    Standard layout: headers in Row 1, template row (formulas or literal
    defaults) in Row 2, data from Row 3. Real-world sheets may have title rows
    above the table; if Row 1 is blank we scan for the first row containing
    2+ non-empty cells (title rows hold a single merged label). For such
    sheets the row after the header is a template only if it holds a formula.
    """
    if ws is None or ws.max_row is None or ws.max_row < 1:
        return EXCEL_HEADER_ROW, True, 1

    if any(c.value not in (None, "") for c in ws[EXCEL_HEADER_ROW]):
        cols = [c.column for c in ws[EXCEL_HEADER_ROW] if c.value not in (None, "")]
        first_col = min(cols) if cols else 1
        return EXCEL_HEADER_ROW, True, first_col

    header_row = None
    for row in range(1, ws.max_row + 1):
        count = sum(1 for c in ws[row] if c.value not in (None, ""))
        if count >= 2:
            header_row = row
            break
    if header_row is None:
        return EXCEL_HEADER_ROW, True, 1

    cols = [c.column for c in ws[header_row] if c.value not in (None, "")]
    first_col = min(cols) if cols else 1

    template_row = header_row + 1
    has_template = False
    if ws.max_row >= template_row:
        has_formula = any(
            c.data_type == "f" or (isinstance(c.value, str) and c.value.startswith("="))
            for c in ws[template_row]
            if c.value not in (None, "")
        )
        has_template = has_formula
    return header_row, has_template, first_col


def _classify(ws, df, headers, template_row, first_col, declared_types=None) -> tuple:
    formula_cols = []
    date_cols = []
    numeric_cols = []
    declared_types = declared_types or {}
    template_row_values = None
    if ws is not None and ws.max_row >= template_row:
        template_row_values = [
            c.value for c in ws[template_row] if not isinstance(c, MergedCell)
        ]

    for col_idx, header in enumerate(headers):
        template_value = None
        if template_row_values is not None and col_idx < len(template_row_values):
            template_value = template_row_values[col_idx]

        is_formula = isinstance(template_value, str) and template_value.startswith("=")
        if not is_formula and ws is not None and ws.max_row >= template_row:
            cell = ws.cell(row=template_row, column=first_col + col_idx)
            is_formula = not isinstance(cell, MergedCell) and cell.data_type == "f"

        col_data = df.iloc[:, col_idx]
        non_empty = [v for v in col_data if v != ""]
        declared_type = declared_types.get(header)
        numeric_count = sum(1 for v in non_empty if _NUMERIC_RE.match(str(v)) is not None)
        if not non_empty:
            if declared_type == "number":
                numeric_cols.append(header)
        elif numeric_count / len(non_empty) >= 0.75:
            numeric_cols.append(header)

        header_upper = header.upper()
        date_cells = sum(1 for v in non_empty if _parse_datetime(str(v)) is not None)
        if "DATE" in header_upper:
            date_cols.append(header)
        elif not non_empty:
            if declared_type == "date":
                date_cols.append(header)
        elif date_cells / len(non_empty) >= 0.75:
            date_cols.append(header)

        if is_formula:
            formula_cols.append(header)

    return formula_cols, numeric_cols, date_cols


def _parse_datetime(value: str):
    """Parse a user-supplied date string into a naive datetime.

    The value is entered by the user and is interpreted as Asia/Kolkata (IST)
    wall-clock time; it never comes from the system clock. The result is kept
    naive because xlsx has no timezone concept and openpyxl rejects tz-aware
    datetimes (the stored wall-clock IS IST)."""
    match = _DATETIME_RE.match(value)
    if not match:
        return None
    year, month, day = int(match.group(1)), int(match.group(2)), int(match.group(3))
    hour = int(match.group(4) or 0)
    minute = int(match.group(5) or 0)
    second = int(match.group(6) or 0)
    try:
        return datetime(year, month, day, hour, minute, second)
    except ValueError:
        return None


def _load_workbook(workbook: str):
    path = config.DATA_DIR / workbook
    if not path.exists():
        raise SheetError(f"workbook not found: {workbook}")
    try:
        wb_obj = openpyxl.load_workbook(path, data_only=False)
    except zipfile.BadZipFile as exc:
        raise SheetError(f"not a valid xlsx file: {workbook}") from exc
    return path, wb_obj


def _build_default_frame():
    return pd.DataFrame(columns=_default_columns)  # type: ignore[arg-type]


def _drop_unnamed_empty_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Drop pandas' fabricated 'Unnamed: N' headers when the column is empty."""
    for col in [c for c in df.columns if str(c).startswith("Unnamed:")]:
        if (df[col].astype(str) == "").all():
            df = df.drop(columns=[col])
    return df


def load_sheet(workbook: str, sheet_name: str) -> Sheet:
    path, wb_obj = _load_workbook(workbook)
    sheet = Sheet(workbook, sheet_name, path, wb_obj, None)

    if sheet_name not in wb_obj.sheetnames:
        df = _build_default_frame()
        sheet.headers = list(_default_columns)
        sheet.orig_types = []
        sheet.formula_cols, sheet.numeric_cols, sheet.date_cols = [], [], []
    else:
        ws = wb_obj[sheet_name]
        sheet.ws = ws
        header_row, has_template, first_col = _detect_layout(ws)
        sheet.header_row = header_row
        sheet.template_row = header_row + 1
        sheet.data_start_row = header_row + (2 if has_template else 1)
        sheet.first_col = first_col

        skiprows = [header_row] if has_template else []
        df = pd.read_excel(
            path, sheet_name=sheet_name, dtype=str, header=header_row - 1, skiprows=skiprows
        ).fillna("")
        df = _drop_unnamed_empty_columns(df)
        if df.empty and df.columns.empty:
            df = _build_default_frame()
        sheet.headers = [str(h) for h in df.columns]
        sheet.template = _capture_template(ws, sheet.template_row) if has_template else {}
        declared_types = settings.get_column_types(workbook, sheet_name)
        sheet.formula_cols, sheet.numeric_cols, sheet.date_cols = _classify(
            ws, df, sheet.headers, sheet.template_row, first_col, declared_types
        )
        sheet.orig_types = _capture_orig_types(ws, len(df), sheet.data_start_row)

    sheet.df = df
    sheet.has_flagged = config.RESERVED_COLUMN in sheet.headers
    return sheet


def _capture_orig_types(ws, row_count: int, data_start_row: int) -> list:
    orig_types = []
    for i in range(row_count):
        excel_row = data_start_row + i
        row_types = {}
        for cell in ws[excel_row]:
            if isinstance(cell, MergedCell):
                continue
            row_types[cell.column] = cell.data_type
        orig_types.append(row_types)
    return orig_types


def list_workbooks():
    if not config.DATA_DIR.exists():
        return []
    result = []
    tz = ZoneInfo(config.TIMEZONE)
    for path in sorted(config.DATA_DIR.glob("*.xlsx")):
        stat = path.stat()
        updated = datetime.fromtimestamp(stat.st_mtime, tz)
        result.append({"name": path.name, "updated_at": updated.strftime("%Y-%m-%d %H:%M:%S")})
    return result


def list_sheets(workbook: str):
    path, wb_obj = _load_workbook(workbook)
    return list(wb_obj.sheetnames)


def _is_identifier_column(header: str) -> bool:
    upper = header.upper()
    return any(k in upper for k in IDENTIFIER_KEYWORDS)


def _to_excel_value(value: str, orig_type, is_date_col: bool, is_identifier: bool):
    if value == "":
        return None
    if is_date_col and orig_type in ("d", NEW_TYPE):
        parsed = _parse_datetime(value)
        if parsed is not None:
            return parsed
    if not is_identifier and _NUMERIC_RE.match(value):
        return float(value.replace(",", "."))
    return value


def save_sheet(sheet: Sheet) -> None:
    ws = sheet.ws
    if ws is None:
        ws = sheet.wb_obj.create_sheet(title=sheet.sheet_name)
        sheet.ws = ws

    max_col = max(ws.max_column, sheet.first_col + len(sheet.headers) - 1)
    if ws.max_row >= sheet.data_start_row:
        for row in ws.iter_rows(min_row=sheet.data_start_row, max_row=ws.max_row,
                                min_col=1, max_col=max_col):
            for c in row:
                if isinstance(c, MergedCell):
                    continue
                c.value = None

    for col in range(sheet.first_col, max_col + 1):
        cell = ws.cell(row=sheet.header_row, column=col)
        if isinstance(cell, MergedCell):
            continue
        cell.value = None

    for col_idx, header in enumerate(sheet.headers):
        ws.cell(row=sheet.header_row, column=sheet.first_col + col_idx, value=header)

    if sheet.template:
        for col_idx, spec in sheet.template.items():
            cell = ws.cell(row=sheet.template_row, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            cell.value = spec.get("value")
            if spec.get("number_format"):
                cell.number_format = spec["number_format"]

    for row_idx, row in enumerate(sheet.df.itertuples(index=False)):
        excel_row = sheet.data_start_row + row_idx
        orig = sheet.orig_types[row_idx] if row_idx < len(sheet.orig_types) else {}
        for col_idx, header in enumerate(sheet.headers):
            raw = row[col_idx]
            value = "" if pd.isna(raw) else str(raw)
            orig_type = orig.get(sheet.first_col + col_idx, None)
            is_date_col = header in sheet.date_cols
            is_identifier = _is_identifier_column(header)
            cell = ws.cell(row=excel_row, column=sheet.first_col + col_idx)
            if isinstance(cell, MergedCell):
                continue
            converted = _to_excel_value(value, orig_type, is_date_col, is_identifier)
            cell.value = converted
            if isinstance(converted, datetime) and not cell.number_format:
                cell.number_format = DATE_NUMBER_FORMAT


def _backup_dir():
    config.BAK_DIR.mkdir(parents=True, exist_ok=True)
    return config.BAK_DIR


def create_backup(workbook: str) -> None:
    src = config.DATA_DIR / workbook
    if not src.exists():
        return
    stamp = datetime.now(ZoneInfo(config.TIMEZONE)).strftime("%Y%m%d-%H%M%S-%f")
    dest = _backup_dir() / f"{workbook}.{stamp}.bak"
    dest.write_bytes(src.read_bytes())

    backups = sorted(_backup_dir().glob(f"{workbook}.*.bak"))
    for old in backups[: -config.BACKUP_KEEP]:
        old.unlink()


def _backup_stamp_dt(stamp: str):
    for fmt in ("%Y%m%d-%H%M%S-%f", "%Y%m%d-%H%M%S"):
        try:
            return datetime.strptime(stamp, fmt)
        except ValueError:
            continue
    return None


def _read_audit_entries() -> list:
    """Parse audit.log into [{ts, action, workbook, sheet, excel_row, description}]."""
    path = config.AUDIT_LOG
    if not path.exists():
        return []
    entries = []
    with open(path, encoding="utf-8") as fh:
        for line in fh:
            parts = line.strip().split(" | ")
            if len(parts) < 5:
                continue
            ts, action, workbook, sheet, excel_row = parts[:5]
            description = " | ".join(parts[5:])
            try:
                ts = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S IST")
            except ValueError:
                try:
                    ts = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    continue
            entries.append({
                "ts": ts,
                "action": action,
                "workbook": workbook,
                "sheet": sheet,
                "excel_row": excel_row,
                "description": description,
            })
    return entries


def list_backups(workbook: str):
    result = []
    backups = sorted(_backup_dir().glob(f"{workbook}.*.bak"), reverse=True)
    entries = [e for e in _read_audit_entries() if e["workbook"] == workbook]
    for path in backups:
        stat = path.stat()
        stamp = path.stem.split(".")[-1]
        created_dt = _backup_stamp_dt(stamp)
        try:
            created = created_dt.strftime("%Y-%m-%d %H:%M:%S") if created_dt else ""
        except ValueError:
            created = ""
        description = f"Version from {created}"
        if created_dt:
            best = None
            best_delta = None
            for entry in entries:
                delta = abs((entry["ts"] - created_dt).total_seconds())
                if best_delta is None or delta < best_delta:
                    best_delta = delta
                    best = entry
            if best is not None and best_delta is not None and best_delta <= 10:
                label = f"{best['action']} — {best['sheet']}"
                if best["excel_row"] not in ("", "0"):
                    label += f" — row {best['excel_row']}"
                description = label
        result.append({
            "filename": path.name,
            "created_at": created,
            "size_bytes": stat.st_size,
            "description": description,
        })
    return result[: config.BACKUP_KEEP]


def backup_path(filename: str):
    if "/" in filename or "\\" in filename or filename.startswith("."):
        raise SheetError("invalid backup filename")
    path = (_backup_dir() / filename).resolve()
    if not str(path).startswith(str(_backup_dir().resolve())):
        raise SheetError("invalid backup path")
    if not path.is_file():
        raise SheetError("backup not found")
    return path


def revert_to_backup(workbook: str, filename: str) -> None:
    path = backup_path(filename)
    if not filename.startswith(workbook + "."):
        raise SheetError("backup does not belong to this workbook")
    current = config.DATA_DIR / workbook
    if not current.exists():
        raise SheetError("workbook not found")
    create_backup(workbook)
    tmp = current.with_suffix(".xlsx.tmp")
    tmp.write_bytes(path.read_bytes())
    tmp.replace(current)
    _audit("REVERT", workbook, "", "", f"restored to {filename}")


def _commit(sheet: Sheet, action: str, excel_row, description: str) -> None:
    save_sheet(sheet)
    tmp = sheet.path.with_suffix(".xlsx.tmp")
    sheet.wb_obj.save(tmp)
    tmp.replace(sheet.path)
    create_backup(sheet.workbook)
    _audit(action, sheet.workbook, sheet.sheet_name, excel_row, description)


def _duplicate_rows(sheet: Sheet, values: dict, exclude_index=None) -> list:
    matches = []
    configured = settings.get_duplicate_columns(sheet.workbook, sheet.sheet_name)
    if configured:
        scan_headers = [h for h in configured if h in sheet.headers and h != config.RESERVED_COLUMN]
    else:
        scan_headers = [h for h in sheet.headers if h != config.RESERVED_COLUMN]
    keyword_filter = not configured
    for header in scan_headers:
        header_upper = header.upper()
        if keyword_filter and not any(k in header_upper for k in DUP_KEYWORDS):
            continue
        if header not in values or values[header] == "":
            continue
        needle = str(values[header]).strip().lower()
        if needle == "":
            continue
        col_idx = sheet.header_index(header)
        for index, row in enumerate(sheet.df.itertuples(index=False)):
            if exclude_index is not None and index == exclude_index:
                continue
            candidate = "" if pd.isna(row[col_idx]) else str(row[col_idx]).strip().lower()
            if candidate == needle:
                matches.append({
                    "excel_row": sheet.excel_row_of(index),
                    "column": header,
                    "value": str(values[header]),
                })
    return matches


def add_row(sheet: Sheet, values: dict, force: bool = False) -> dict:
    for header in sheet.numeric_cols:
        if header in values and str(values.get(header, "")).strip() == "":
            values[header] = "0"

    new_row = {h: values.get(h, "") for h in sheet.headers if h != config.RESERVED_COLUMN}
    if sheet.has_flagged:
        new_row[config.RESERVED_COLUMN] = ""

    duplicates = _duplicate_rows(sheet, new_row)
    if duplicates and not force:
        raise DuplicateError(duplicates)

    top = settings.get_append_direction(sheet.workbook) == "top"
    new_excel_row = sheet.data_start_row if top else sheet.row_count + sheet.data_start_row
    computed, formula_warnings = compute_formula_values(
        sheet, new_row, "top" if top else "bottom", new_excel_row
    )
    for header, value in computed.items():
        new_row[header] = value
    sheet.warnings = formula_warnings

    frame = pd.DataFrame([new_row], columns=sheet.headers)  # type: ignore[arg-type]
    if top:
        sheet.df = pd.concat([frame, sheet.df], ignore_index=True)
        new_index = 0
    else:
        sheet.df = pd.concat([sheet.df, frame], ignore_index=True)
        new_index = len(sheet.df) - 1
    sheet.orig_types.insert(new_index, {c: NEW_TYPE for c in range(1, len(sheet.headers) + 1)})

    excel_row = sheet.excel_row_of(new_index)
    _commit(sheet, "ADD", excel_row, _redacted(new_row, sheet.headers))
    return {
        "excel_row": excel_row,
        "appended_to": "top" if top else "bottom",
        "warnings": formula_warnings,
        "duplicates": duplicates,
    }


def update_row(sheet: Sheet, excel_row: int, values: dict, force: bool = False) -> dict:
    index = excel_row - sheet.data_start_row
    if index < 0 or index >= len(sheet.df):
        raise SheetError("row out of range")

    new_values = {h: values.get(h, "") for h in sheet.headers if h != config.RESERVED_COLUMN}
    for header in sheet.formula_cols:
        new_values[header] = sheet.cell_value(header, index)

    duplicates = _duplicate_rows(sheet, new_values, exclude_index=index)
    if duplicates and not force:
        raise DuplicateError(duplicates)

    computed, formula_warnings = compute_formula_values(sheet, new_values, "edit", excel_row)
    for header, value in computed.items():
        new_values[header] = value
    sheet.warnings = formula_warnings

    description = _describe_update(sheet, index, new_values)
    for header, value in new_values.items():
        sheet.set_cell(header, index, value)

    _commit(sheet, "EDIT", excel_row, description)
    return {"excel_row": excel_row, "warnings": formula_warnings, "duplicates": duplicates}


def _describe_update(sheet: Sheet, index: int, new_values: dict) -> str:
    parts = []
    for header in sheet.headers:
        if header == config.RESERVED_COLUMN:
            continue
        old = sheet.cell_value(header, index)
        new = str(new_values.get(header, ""))
        if old == new:
            continue
        lowered = header.upper()
        if any(k in lowered for k in config.BANK_ACCOUNT_KEYWORDS):
            parts.append(f"{header} <changed>")
        else:
            parts.append(f"{header} {old} -> {new}")
    return ", ".join(parts) if parts else "no change"


def delete_row(sheet: Sheet, excel_row: int) -> None:
    index = excel_row - sheet.data_start_row
    if index < 0 or index >= len(sheet.df):
        raise SheetError("row out of range")
    description = _redacted(
        {h: sheet.cell_value(h, index) for h in sheet.headers},
        sheet.headers,
    )
    sheet.df = sheet.df.drop(index=index).reset_index(drop=True)
    if index < len(sheet.orig_types):
        sheet.orig_types.pop(index)
    _commit(sheet, "DELETE", excel_row, description)


def toggle_flag(sheet: Sheet, excel_row: int) -> dict:
    index = excel_row - sheet.data_start_row
    if index < 0 or index >= len(sheet.df):
        raise SheetError("row out of range")

    flagged_column_added = False
    if not sheet.has_flagged:
        sheet.headers.append(config.RESERVED_COLUMN)
        sheet.df[config.RESERVED_COLUMN] = ""
        sheet.has_flagged = True
        flagged_column_added = True
        for row_types in sheet.orig_types:
            row_types[len(sheet.headers)] = "s"

    current = sheet.cell_value(config.RESERVED_COLUMN, index)
    new_value = "" if current.upper() == "TRUE" else "TRUE"
    sheet.set_cell(config.RESERVED_COLUMN, index, new_value)
    flagged = new_value == "TRUE"

    action = "FLAG" if flagged else "UNFLAG"
    sheet.warnings = []
    _commit(sheet, action, excel_row, "flagged" if flagged else "unflagged")
    return {"flagged": flagged, "flagged_column_added": flagged_column_added}


def _normalize_columns(columns) -> list:
    """Validate/normalize a user-supplied column list.

    Accepts [{"name": str, "type": "text"|"number"|"date"}, ...]. Returns a
    list of (name, type) tuples. Raises SheetError on a bad shape."""
    if not isinstance(columns, list):
        raise SheetError("columns must be a list")
    result = []
    seen = set()
    for entry in columns:
        if not isinstance(entry, dict):
            raise SheetError("each column must be an object with name and type")
        name = entry.get("name", "")
        col_type = entry.get("type", "text")
        if not isinstance(name, str) or not name.strip():
            raise SheetError("column name must be a non-empty string")
        name = name.strip()
        if name == config.RESERVED_COLUMN:
            raise SheetError(f"reserved column name: {name}")
        if name in seen:
            raise SheetError(f"duplicate column: {name}")
        seen.add(name)
        if col_type not in ("text", "number", "date"):
            raise SheetError(f"invalid column type for {name}: {col_type}")
        result.append((name, col_type))
    return result


def add_sheet(workbook: str, sheet_name: str, columns=None) -> None:
    path, wb_obj = _load_workbook(workbook)
    if sheet_name in wb_obj.sheetnames:
        raise SheetError(f"sheet already exists: {sheet_name}")
    normalized = None
    if columns is not None:
        normalized = _normalize_columns(columns)
    headers = _default_columns if normalized is None else [name for name, _ in normalized]
    ws = wb_obj.create_sheet(title=sheet_name)
    for col_idx, header in enumerate(headers):
        ws.cell(row=EXCEL_HEADER_ROW, column=col_idx + 1, value=header)
    tmp = path.with_suffix(".xlsx.tmp")
    wb_obj.save(tmp)
    tmp.replace(path)
    create_backup(workbook)
    _audit("ADD_SHEET", workbook, sheet_name, "", "added sheet" + (" with columns" if normalized else ""))
    if normalized:
        settings.set_column_types(workbook, sheet_name, dict(normalized))


_file_locks_guard = threading.Lock()
_file_locks: dict = {}


def file_lock(workbook: str) -> threading.Lock:
    with _file_locks_guard:
        if workbook not in _file_locks:
            _file_locks[workbook] = threading.Lock()
        return _file_locks[workbook]


def compute_formula_values(sheet, row_values, insert_mode, new_excel_row):
    from formula import compute_formula_values as _impl
    return _impl(sheet, row_values, insert_mode, new_excel_row)


def _template_formula_values(sheet: Sheet) -> dict:
    result = {}
    for col_idx, header in enumerate(sheet.headers):
        spec = sheet.template.get(sheet.first_col + col_idx, {})
        value = spec.get("value")
        result[header] = value if isinstance(value, str) and value.startswith("=") else ""
    return result


def template_formulas(sheet: Sheet) -> dict:
    """Read every column's template-row formula plus its Excel column letter.

    Returns {header: {"formula": formula_text, "ref": column_letter}}."""
    result = {}
    for col_idx, header in enumerate(sheet.headers):
        spec = sheet.template.get(sheet.first_col + col_idx, {})
        value = spec.get("value")
        formula = value if isinstance(value, str) and value.startswith("=") else ""
        letter = openpyxl.utils.get_column_letter(sheet.first_col + col_idx)  # type: ignore[attr-defined]
        result[header] = {"formula": formula, "ref": letter}
    return result
    return result


def _add_column(sheet: Sheet, header: str) -> None:
    sheet.headers.append(header)
    if header not in sheet.df.columns:
        sheet.df[header] = ""
    col_abs = sheet.first_col + sheet.header_index(header)
    for row_types in sheet.orig_types:
        row_types[col_abs] = "s"


def _recompute_rows(sheet: Sheet) -> list:
    warnings = []
    for index in range(len(sheet.df)):
        excel_row = sheet.data_start_row + index
        row_values = {h: sheet.cell_value(h, index) for h in sheet.headers}
        computed, row_warnings = compute_formula_values(sheet, row_values, "edit", excel_row)
        for header, value in computed.items():
            sheet.set_cell(header, index, value)
        warnings.extend(row_warnings)
    return warnings


def update_formulas(sheet: Sheet, formulas: dict) -> dict:
    from formula import validate_formula

    submitted = {}
    for header, formula_text in (formulas or {}).items():
        submitted[str(header)] = "" if formula_text is None else str(formula_text).strip()

    errors = {}
    for header, text in submitted.items():
        if text == "":
            continue
        ok, reason = validate_formula(sheet, header, text)
        if not ok:
            errors[header] = reason
    if errors:
        raise FormulaValidationError(errors)

    description_parts = []
    for header, text in submitted.items():
        if header not in sheet.headers:
            _add_column(sheet, header)
        col_abs = sheet.first_col + sheet.header_index(header)
        if text == "":
            sheet.template.pop(col_abs, None)
            if header in sheet.formula_cols:
                sheet.formula_cols.remove(header)
        else:
            sheet.template[col_abs] = {"value": text, "data_type": "f", "number_format": None}
            if header not in sheet.formula_cols:
                sheet.formula_cols.append(header)
        description_parts.append(f"{header}={text}")

    warnings = []
    if sheet.formula_cols and len(sheet.df):
        warnings = _recompute_rows(sheet)

    description = ", ".join(description_parts) if description_parts else "no change"
    _commit(sheet, "EDIT_FORMULA", "", description)
    return {"formulas": _template_formula_values(sheet), "warnings": warnings}


def add_column(sheet: Sheet, header: str, col_type: str) -> None:
    if header in sheet.headers:
        raise SheetError(f"column already exists: {header}")
    if header == config.RESERVED_COLUMN:
        raise SheetError(f"reserved column name: {header}")
    if col_type not in ("text", "number", "date"):
        raise SheetError(f"invalid column type: {col_type}")
    sheet.headers.append(header)
    sheet.df[header] = ""
    col_abs = sheet.first_col + sheet.header_index(header)
    for row_types in sheet.orig_types:
        row_types[col_abs] = "s"
    if col_type == "number":
        sheet.numeric_cols.append(header)
    elif col_type == "date":
        sheet.date_cols.append(header)
    settings.set_column_type(sheet.workbook, sheet.sheet_name, header, col_type)
    _commit(sheet, "ADD_COLUMN", "", f"added column {header} ({col_type})")


def delete_column(sheet: Sheet, header: str) -> None:
    if header not in sheet.headers:
        raise SheetError(f"unknown column: {header}")
    if header == config.RESERVED_COLUMN:
        raise SheetError("cannot delete the flagged column")
    col_idx = sheet.header_index(header)
    col_abs = sheet.first_col + col_idx
    sheet.df = sheet.df.drop(columns=[header])
    sheet.headers.remove(header)
    if header in sheet.formula_cols:
        sheet.formula_cols.remove(header)
    if header in sheet.numeric_cols:
        sheet.numeric_cols.remove(header)
    if header in sheet.date_cols:
        sheet.date_cols.remove(header)
    sheet.template.pop(col_abs, None)
    for row_types in sheet.orig_types:
        shifted = {}
        for c, t in row_types.items():
            if c == col_abs:
                continue
            shifted[c - 1 if c > col_abs else c] = t
        row_types.clear()
        row_types.update(shifted)
    settings.remove_column_type(sheet.workbook, sheet.sheet_name, header)
    dup = settings.get_duplicate_columns(sheet.workbook, sheet.sheet_name)
    if header in dup:
        settings.set_duplicate_columns(
            sheet.workbook, sheet.sheet_name, [c for c in dup if c != header]
        )
    _commit(sheet, "DELETE_COLUMN", "", f"deleted column {header}")


def rename_column(sheet: Sheet, old_name: str, new_name: str) -> None:
    if old_name not in sheet.headers:
        raise SheetError(f"unknown column: {old_name}")
    if old_name == config.RESERVED_COLUMN:
        raise SheetError("cannot rename the flagged column")
    if not isinstance(new_name, str) or not new_name.strip():
        raise SheetError("column name must be a non-empty string")
    new_name = new_name.strip()
    if new_name == config.RESERVED_COLUMN:
        raise SheetError(f"reserved column name: {new_name}")
    if new_name in sheet.headers:
        raise SheetError(f"column already exists: {new_name}")
    sheet.headers[sheet.header_index(old_name)] = new_name
    sheet.df = sheet.df.rename(columns={old_name: new_name})
    for kind in (sheet.formula_cols, sheet.numeric_cols, sheet.date_cols):
        if old_name in kind:
            kind.remove(old_name)
            kind.append(new_name)
    settings.rename_column_type(sheet.workbook, sheet.sheet_name, old_name, new_name)
    dup = settings.get_duplicate_columns(sheet.workbook, sheet.sheet_name)
    if old_name in dup:
        settings.set_duplicate_columns(
            sheet.workbook, sheet.sheet_name, [new_name if c == old_name else c for c in dup]
        )
    _commit(sheet, "RENAME_COLUMN", "", f"{old_name} -> {new_name}")